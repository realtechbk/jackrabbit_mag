import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

OUT = sys.argv[1]

DARK, PURPLE, LABELBG, CALCBG, WHITE = "2F3641", "8B3DD7", "D6DEF5", "DCE3F2", "FFFFFF"
AMBER = "FFD479"
F = "Arial"

w   = Font(name=F, size=10, color=WHITE)
wb_ = Font(name=F, size=10, bold=True, color=WHITE)
hdr = Font(name=F, size=10, bold=True, color=WHITE)
inp = Font(name=F, size=10, color="0000FF")
cal = Font(name=F, size=10, color="000000")
tot = Font(name=F, size=10, bold=True, color="000000")

dark  = PatternFill("solid", fgColor=DARK)
purp  = PatternFill("solid", fgColor=PURPLE)
lab   = PatternFill("solid", fgColor=LABELBG)
calc  = PatternFill("solid", fgColor=CALCBG)
white = PatternFill("solid", fgColor=WHITE)

thin = Side(style="thin", color="9AA3B2")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

MONEY = '$#,##0.00;($#,##0.00);-'
NUM   = '#,##0;(#,##0);-'

ROWS = [
    "Recreational Tuition -- Preschool (The Jungle)",
    "Recreational Tuition -- Girls Wings",
    "Recreational Tuition -- Boys Wings",
    "Recreational Tuition -- Tumbling",
    "Competitive -- Pre-Team",
    "Competitive -- American Flyers Teams",
    "Ancillary -- Open Gym",
    "Ancillary -- Annual Membership Fees",
    "Events -- American Flyers Cup",
]

wbk = Workbook()


def canvas(ws, nrows, ncols):
    ws.sheet_view.showGridLines = False
    for r in range(1, nrows + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = dark
            cell.font = w


def header(ws, row, col, text, width=None):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = purp
    c.font = hdr
    c.border = box
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if width:
        ws.column_dimensions[c.column_letter].width = width
    return c


# ============ Instructions ============
ins = wbk.active
ins.title = "Instructions"
canvas(ins, 60, 8)
ins.column_dimensions['A'].width = 2
ins.column_dimensions['B'].width = 30
ins.column_dimensions['C'].width = 96

ins['B1'] = "Jackrabbit → Service Structure Crosswalk"
ins['B1'].font = Font(name=F, size=16, bold=True, color=WHITE)
ins['B2'] = "Maine Academy of Gymnastics · check that Jackrabbit's data actually separates into the nine service rows"
ins['B2'].font = Font(name=F, size=10, italic=True, color="C9D1E0")

steps = [
    ("STEP 1", "Export the class list.", True),
    ("", "Classes → Reports → Class Listing.  Select these columns before exporting:", False),
    ("", "Class Name · Category 1 · Category 2 · Category 3 · Tuition/Fee · Instructor · Day · Time ·", False),
    ("", "Duration · Max Enrollment · Current Enrollment · Session · Status", False),
    ("", "Set the filter to include INACTIVE classes as well — 2023 and 2024 classes may be archived.", False),
    ("", "", False),
    ("STEP 2", "Paste it into the 'Class Export' tab, starting at cell A5.", True),
    ("", "Match the column order shown there. Columns N and O will populate automatically.", False),
    ("", "", False),
    ("STEP 3", "List the distinct Category 1 values on the 'Category Crosswalk' tab.", True),
    ("", "Gear icon → Settings → Drop-down Lists → Class Category 1 gives you the full list.", False),
    ("", "Map each one to a service row using the drop-down in column B.", False),
    ("", "", False),
    ("STEP 4", "List the transaction sub-types on the 'Subtype Crosswalk' tab.", True),
    ("", "Gear icon → Settings → Drop-down Lists → Transaction Sub-type.", False),
    ("", "This is what determines whether tuition, the $45 membership fee, Open Gym and meet fees", False),
    ("", "can be separated in the ledger at all.", False),
    ("", "", False),
    ("STEP 5", "Read the flag summary at the top of the 'Class Export' tab.", True),
    ("", "Anything not showing OK is either a Jackrabbit cleanup task or a documented estimate.", False),
]
r = 4
for tag, text, bold in steps:
    if tag:
        c = ins.cell(row=r, column=2, value=tag)
        c.font = Font(name=F, size=10, bold=True, color=AMBER)
    c2 = ins.cell(row=r, column=3, value=text)
    c2.font = Font(name=F, size=10, bold=bold, color=WHITE)
    r += 1

r += 1
ins.cell(row=r, column=2, value="WHAT YOU ARE LOOKING FOR").font = Font(name=F, size=11, bold=True, color=WHITE)
r += 1
findings = [
    "Unmapped classes — blank Category 1, or a category that fits none of the nine rows.",
    "Collisions — one Jackrabbit category spanning two service rows (e.g. a single 'Rec' category covering both",
    "        girls' and boys' Wings). If so, the split has to come from class names instead of categories.",
    "Blended sub-types — membership fees or Open Gym posting as generic 'Tuition', which makes those rows impossible to isolate.",
    "Category drift — classes re-tagged between 2023 and 2025. This breaks year-over-year comparability.",
]
for text in findings:
    ins.cell(row=r, column=3, value=text).font = w
    r += 1

r += 2
ins.cell(row=r, column=2, value="Do this BEFORE pulling any revenue numbers.").font = Font(name=F, size=10, bold=True, color=AMBER)
ins.cell(row=r, column=3, value="The crosswalk determines whether a revenue pull is even valid.").font = w

# ============ Class Export ============
ce = wbk.create_sheet("Class Export")
canvas(ce, 320, 17)
ce.column_dimensions['A'].width = 2

ce['B1'] = "Class Export — paste the Jackrabbit Class Listing below, starting at B5"
ce['B1'].font = Font(name=F, size=13, bold=True, color=WHITE)

FR, LR = 5, 304

# summary block
sums = [
    ("Classes pasted",                 f'=COUNTA(B{FR}:B{LR})'),
    ("Mapped OK",                      f'=COUNTIF(P{FR}:P{LR},"OK")'),
    ("No Category 1",                  f'=COUNTIF(P{FR}:P{LR},"NO CATEGORY 1")'),
    ("Category not in crosswalk",      f'=COUNTIF(P{FR}:P{LR},"CATEGORY NOT IN CROSSWALK")'),
    ("No tuition on class record",     f'=COUNTIF(P{FR}:P{LR},"NO TUITION")'),
]
for i, (label, formula) in enumerate(sums):
    c1 = ce.cell(row=1 + (i % 3), column=8 + 3 * (i // 3), value=label)
    c1.font = wb_
    c1.alignment = Alignment(horizontal="right")
    c2 = ce.cell(row=1 + (i % 3), column=9 + 3 * (i // 3), value=formula)
    c2.fill = white
    c2.font = tot
    c2.number_format = NUM
    c2.border = box
    c2.alignment = Alignment(horizontal="center")

cols = [
    ("Class Name", 30), ("Category 1", 18), ("Category 2", 14), ("Category 3", 14),
    ("Tuition / Fee", 13), ("Instructor", 16), ("Day", 10), ("Time", 10),
    ("Duration", 10), ("Max Enroll", 10), ("Current Enroll", 12), ("Session", 14),
    ("Status", 12), ("Maps To (auto)", 34), ("Flag (auto)", 26),
]
for i, (name, width) in enumerate(cols):
    header(ce, 4, 2 + i, name, width)

CAT_A = "'Category Crosswalk'!$B$6:$B$45"
CAT_B = "'Category Crosswalk'!$C$6:$C$45"

for r in range(FR, LR + 1):
    for c in range(2, 15):
        cell = ce.cell(row=r, column=c)
        cell.fill = white
        cell.font = inp
        cell.border = box
        if c == 6:
            cell.number_format = MONEY
        elif c in (11, 12):
            cell.number_format = NUM

    maps = ce.cell(row=r, column=15,
                   value=(f'=IF($C{r}="","",'
                          f'IFERROR(INDEX({CAT_B},MATCH($C{r},{CAT_A},0)),"UNMAPPED"))'))
    maps.fill = calc
    maps.font = cal
    maps.border = box
    maps.alignment = Alignment(horizontal="left", indent=1)

    flag = ce.cell(row=r, column=16,
                   value=(f'=IF($B{r}="","",'
                          f'IF($C{r}="","NO CATEGORY 1",'
                          f'IF(ISNA(MATCH($C{r},{CAT_A},0)),"CATEGORY NOT IN CROSSWALK",'
                          f'IF($F{r}="","NO TUITION","OK"))))'))
    flag.fill = calc
    flag.font = cal
    flag.border = box
    flag.alignment = Alignment(horizontal="center")

ce.freeze_panes = "C5"

# ============ Category Crosswalk ============
cc = wbk.create_sheet("Category Crosswalk")
canvas(cc, 60, 10)
cc.column_dimensions['A'].width = 2
cc['B1'] = "Category Crosswalk — map every Jackrabbit Category 1 value to one service row"
cc['B1'].font = Font(name=F, size=13, bold=True, color=WHITE)
cc['B2'] = "Gear icon → Settings → Drop-down Lists → Class Category 1"
cc['B2'].font = Font(name=F, size=10, italic=True, color="C9D1E0")

header(cc, 5, 2, "Jackrabbit Category 1 value", 34)
header(cc, 5, 3, "Maps to Revenue Group -- Sub Group", 40)
header(cc, 5, 4, "Notes / issue", 46)

for r in range(6, 46):
    for c in (2, 3, 4):
        cell = cc.cell(row=r, column=c)
        cell.fill = white
        cell.font = inp
        cell.border = box
        cell.alignment = Alignment(horizontal="left", indent=1)

# reference list for validation
cc.cell(row=5, column=8, value="Valid service rows").font = Font(name=F, size=9, bold=True, color=AMBER)
for i, name in enumerate(ROWS):
    c = cc.cell(row=6 + i, column=8, value=name)
    c.fill = lab
    c.font = Font(name=F, size=9, color="1F2937")
    c.border = box
cc.column_dimensions['H'].width = 42

dv = DataValidation(type="list", formula1="='Category Crosswalk'!$H$6:$H$14", allow_blank=True)
dv.error = "Pick one of the nine service rows."
dv.errorTitle = "Not a valid service row"
cc.add_data_validation(dv)
dv.add(f"C6:C45")

# ============ Subtype Crosswalk ============
sc = wbk.create_sheet("Subtype Crosswalk")
canvas(sc, 50, 8)
sc.column_dimensions['A'].width = 2
sc['B1'] = "Transaction Sub-type Crosswalk"
sc['B1'].font = Font(name=F, size=13, bold=True, color=WHITE)
sc['B2'] = "Gear icon → Settings → Drop-down Lists → Transaction Sub-type"
sc['B2'].font = Font(name=F, size=10, italic=True, color="C9D1E0")
sc['B3'] = "If a sub-type covers more than one service row, note it — that row cannot be isolated without cleanup."
sc['B3'].font = Font(name=F, size=10, color=AMBER)

header(sc, 5, 2, "Transaction Sub-type in Jackrabbit", 36)
header(sc, 5, 3, "Belongs to which service row(s)", 40)
header(sc, 5, 4, "Clean or blended?", 20)
header(sc, 5, 5, "Notes", 44)

for r in range(6, 41):
    for c in (2, 3, 4, 5):
        cell = sc.cell(row=r, column=c)
        cell.fill = white
        cell.font = inp
        cell.border = box
        cell.alignment = Alignment(horizontal="left", indent=1)

dv2 = DataValidation(type="list", formula1='"Clean,Blended,Unused,Unclear"', allow_blank=True)
sc.add_data_validation(dv2)
dv2.add("D6:D40")

sc.cell(row=43, column=2, value="Sub-types you should expect to find, if the setup is clean:").font = wb_
expected = [
    "Tuition — recreational",
    "Tuition — team",
    "Annual Membership Fee ($45)",
    "Open Gym",
    "Meet / competition fees",
    "Registration fee",
]
for i, e in enumerate(expected):
    sc.cell(row=44 + i, column=3, value="· " + e).font = w

for ws_ in (ins, ce, cc, sc):
    ws_.page_setup.orientation = "landscape"
    ws_.page_setup.fitToWidth = 1
    ws_.page_setup.fitToHeight = 0
    ws_.sheet_properties.pageSetUpPr.fitToPage = True

wbk.save(OUT)
print("wrote", OUT)
