import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = sys.argv[1]

DARK, PURPLE, LABELBG, CALCBG, WHITE = "2F3641", "8B3DD7", "D6DEF5", "DCE3F2", "FFFFFF"
AMBER, RED = "FFD479", "FF9B9B"
F = "Arial"

w    = Font(name=F, size=10, color=WHITE)
wbf  = Font(name=F, size=10, bold=True, color=WHITE)
hdrf = Font(name=F, size=10, bold=True, color=WHITE)
inp  = Font(name=F, size=10, color="0000FF")
cal  = Font(name=F, size=10, color="000000")
tot  = Font(name=F, size=10, bold=True, color="000000")

dark  = PatternFill("solid", fgColor=DARK)
purp  = PatternFill("solid", fgColor=PURPLE)
lab   = PatternFill("solid", fgColor=LABELBG)
calc  = PatternFill("solid", fgColor=CALCBG)
white = PatternFill("solid", fgColor=WHITE)
totf  = PatternFill("solid", fgColor="B9C4DC")

thin = Side(style="thin", color="9AA3B2")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

MONEY = '$#,##0;($#,##0);-'
MONEY2 = '$#,##0.00;($#,##0.00);-'
PCT   = '0.0%;(0.0%);-'

# ---- 2023 actuals, straight from the Revenue Summary PDF ----
# (row label, group, net revenue 2023, discount 2023, source note)
DATA = [
    ("Recreational Tuition -- Preschool (The Jungle)", "Rev",
     198842.10 + 144493.16, 16429.20 + 14881.79,
     "Cat2 Pre-School $198,842.10 + Parent-Child $144,493.16"),
    ("Recreational Tuition -- Girls Wings", "Rev",
     748653.02 - 162.00, 51277.50,
     "Cat2 Rec Girls $748,653.02 less $162.00 credits"),
    ("Recreational Tuition -- Boys Wings", "Rev",
     86323.49, 9521.98,
     "Cat2 Rec Boys"),
    ("Recreational Tuition -- Tumbling", "Rev",
     83339.09 - 32.00, 10038.11,
     "Cat2 Tumble (Combined) less $32.00 credits"),
    ("Competitive -- Pre-Team", "Rev",
     26248.25, 1947.05,
     "Cat2 Pre-Team (Invite Only) - a real category, not rounding error"),
    ("Competitive -- Team Tuition", "Rev",
     377627.07, 21344.46,
     "Sum of the 10 Team billing-group activities in Cat1 Team"),
    ("Competitive -- Meet Entry Fees", "Rev",
     140104.38, 924.60,
     "Cat1 Team residual: meet entries, AF Cup, unnamed activity. LARGELY PASS-THROUGH"),
    ("Ancillary -- Annual Membership Fees", "Rev",
     58293.50, 0.00,
     "Cat1 Annual Membership"),
    ("Ancillary -- Birthday Parties", "Rev",
     45767.05, 0.00,
     "Cat1 Birthday Parties - NOT on the website navigation"),
    ("Ancillary -- Open Gym", "Rev",
     30059.50, 30.00,
     "Cat1 Open Gym"),
    ("Ancillary -- Pro Shop", "Rev",
     28031.65, 321.31,
     "Cat1 ProShop"),
    ("Ancillary -- Private Lessons", "Rev",
     3815.71, 0.00,
     "Cat1 Private Lessons"),
    ("Ancillary -- Gift Certificates", "Rev",
     310.00, 0.00,
     "Cat1 Gift Certificate"),
    ("Unclassified -- Rec, no Cat2", "Flag",
     21744.45, 1686.40,
     "Cat1 Recreational with blank Cat2 - NEEDS INVESTIGATION"),
    ("EXCLUDE -- Staff", "Excl",
     2660.10, 0.00,
     "Cat1 Staff - not customer revenue"),
    ("EXCLUDE -- Unapplied Payments", "Excl",
     172.00, 0.00,
     "Cat1 --Unapplied Payments--"),
]

GRAND = 1996290.52
GRAND_DISC = 128402.40

wbk = Workbook()


def canvas(ws, nr, nc):
    ws.sheet_view.showGridLines = False
    for r in range(1, nr + 1):
        for c in range(1, nc + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = dark
            cell.font = w


def H(ws, row, col, text, width=None, align="center"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = purp
    c.font = hdrf
    c.border = box
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True, indent=1 if align == "left" else 0)
    if width:
        ws.column_dimensions[c.column_letter].width = width
    return c


# ==================== 2023 Actuals ====================
ws = wbk.active
ws.title = "2023 Actuals"
canvas(ws, 46, 9)
ws.column_dimensions['A'].width = 2

ws['B1'] = "Maine Academy of Gymnastics — 2023 Revenue by Service Line"
ws['B1'].font = Font(name=F, size=15, bold=True, color=WHITE)
ws['B2'] = "Source: Jackrabbit Revenue Summary, 1/1/2023 - 12/31/2023, run 28 July 2026. Reconciles to the report's stated total."
ws['B2'].font = Font(name=F, size=10, italic=True, color="C9D1E0")

H(ws, 4, 2, "Service Row", 44, "left")
H(ws, 4, 3, "Net Revenue 2023", 16)
H(ws, 4, 4, "% of Total", 11)
H(ws, 4, 5, "Discounts Given", 15)
H(ws, 4, 6, "Discount Rate", 12)
H(ws, 4, 7, "Where it comes from", 62, "left")

FR = 5
r = FR
for label, grp, rev, disc, note in DATA:
    lc = ws.cell(row=r, column=2, value=label)
    lc.fill = lab if grp == "Rev" else PatternFill("solid", fgColor="E8D9F5" if grp == "Flag" else "D9D9D9")
    lc.font = Font(name=F, size=10, bold=True, color="1F2937")
    lc.border = box
    lc.alignment = Alignment(horizontal="left", indent=1)

    c = ws.cell(row=r, column=3, value=rev)
    c.fill = white; c.font = inp; c.number_format = MONEY2; c.border = box
    c.alignment = Alignment(horizontal="center")

    c = ws.cell(row=r, column=4, value=f"=IFERROR(C{r}/$C${FR + len(DATA)},0)")
    c.fill = calc; c.font = cal; c.number_format = PCT; c.border = box
    c.alignment = Alignment(horizontal="center")

    c = ws.cell(row=r, column=5, value=disc)
    c.fill = white; c.font = inp; c.number_format = MONEY2; c.border = box
    c.alignment = Alignment(horizontal="center")

    c = ws.cell(row=r, column=6, value=f"=IFERROR(E{r}/(C{r}+E{r}),0)")
    c.fill = calc; c.font = cal; c.number_format = PCT; c.border = box
    c.alignment = Alignment(horizontal="center")

    c = ws.cell(row=r, column=7, value=note)
    c.fill = calc; c.font = Font(name=F, size=9, color="1F2937"); c.border = box
    c.alignment = Alignment(horizontal="left", indent=1)
    r += 1

LR = r - 1
TOTR = r
tl = ws.cell(row=TOTR, column=2, value="TOTAL — reconciles to report")
tl.fill = purp; tl.font = hdrf; tl.border = box
tl.alignment = Alignment(horizontal="left", indent=1)
for col, formula, fmt in ((3, f"=SUM(C{FR}:C{LR})", MONEY2), (4, f"=IFERROR(C{TOTR}/C{TOTR},0)", PCT),
                          (5, f"=SUM(E{FR}:E{LR})", MONEY2), (6, f"=IFERROR(E{TOTR}/(C{TOTR}+E{TOTR}),0)", PCT)):
    c = ws.cell(row=TOTR, column=col, value=formula)
    c.fill = totf; c.font = tot; c.number_format = fmt; c.border = box
    c.alignment = Alignment(horizontal="center")

r = TOTR + 2
checks = [
    ("RECONCILIATION CHECK", Font(name=F, size=11, bold=True, color=WHITE)),
]
for text, font in checks:
    ws.cell(row=r, column=2, value=text).font = font
    r += 1
for label, val, formula in (
    ("Report stated total revenue", GRAND, None),
    ("Sum of rows above", None, f"=C{TOTR}"),
    ("Variance (must be zero)", None, f"=C{TOTR}-C{r}"),
):
    lc = ws.cell(row=r, column=2, value=label)
    lc.font = wbf
    lc.alignment = Alignment(horizontal="right")
    c = ws.cell(row=r, column=3, value=val if val is not None else formula)
    c.fill = white; c.font = tot; c.number_format = MONEY2; c.border = box
    c.alignment = Alignment(horizontal="center")
    r += 1

r += 1
notes = [
    ("MATERIAL FINDINGS", Font(name=F, size=11, bold=True, color=AMBER)),
    ("Four revenue lines exist that the website never mentions: Birthday Parties $45,767 · Pro Shop $28,032 ·", w),
    ("Private Lessons $3,816 · Gift Certificates $310. Together $77,925, or 3.9% of revenue. Birthday Parties alone", w),
    ("is larger than Open Gym, yet there is no Parties page in the site navigation.", w),
    ("", w),
    ("Pre-Team is a real category worth $26,248. I previously recommended dropping this row based on the 2026", w),
    ("class list showing only two classes. That was wrong — 'Pre-Team (Invite Only)' is its own Cat2 with", w),
    ("meaningful revenue. The row stays.", w),
    ("", w),
    ("Meet entry fees are roughly $140k inside Cat1 Team and are largely PASS-THROUGH to host gyms. Leaving them", w),
    ("blended into team revenue would make team gross margin meaningless. Split out as its own row and expect a", w),
    ("near-zero margin on it. Note American Flyers Cup 2023 + 2024 total $9,445 and DO sit on the gym's books.", w),
    ("", w),
    ("Discounts total $128,402 — 6.0% of gross. Recreational runs 7.5%, Team 4.0%. Material enough that gross", w),
    ("versus net revenue must be stated explicitly in any comparison.", w),
    ("", w),
    ("$21,744 of Recreational revenue has NO Cat2 and cannot be attributed to a service. 1.1% of total.", Font(name=F, size=10, bold=True, color=AMBER)),
    ("Almost all of it sits under a single '--No Activity Name--' line. Needs investigation before the trend is trusted.", w),
    ("", w),
    ("THIS REPORT DOES NOT GIVE TRANSACTION COUNTS.", Font(name=F, size=10, bold=True, color=AMBER)),
    ("The Enrollment column reads 0 on every line. For the '# Sales Transactions' and 'Avg. Sales Transaction Value'", w),
    ("columns you still need the Transaction Listing or Transaction Summary report for the same date range.", w),
    ("", w),
    ("Summer Intensives cannot be isolated from this report — they sit inside the Rec Cat2 values. Separating them", w),
    ("requires filtering by session, which the Revenue Summary does not expose. Consider dropping that row.", w),
]
for text, font in notes:
    ws.cell(row=r, column=2, value=text).font = font
    r += 1

# ==================== Team Detail ====================
td = wbk.create_sheet("Team Detail 2023")
canvas(td, 40, 6)
td.column_dimensions['A'].width = 2
td['B1'] = "Cat1 Team — splitting tuition from meet fees"
td['B1'].font = Font(name=F, size=14, bold=True, color=WHITE)
td['B2'] = "Total Cat1 Team revenue 2023: $517,731.45"
td['B2'].font = Font(name=F, size=10, italic=True, color="C9D1E0")

H(td, 4, 2, "Billing-group activity (tuition)", 40, "left")
H(td, 4, 3, "Net Revenue", 15)
H(td, 4, 4, "Discount", 13)

TEAM = [
    ("Team Billing Group 12 hrs", 110220.74, 6871.33),
    ("Team Billing Group 16 hrs", 75080.37, 6014.00),
    ("Team Group 12 hrs", 56482.23, 569.33),
    ("Team Billing Group 20 hrs", 44909.80, 5104.20),
    ("Team Billing 10 hrs", 30033.53, 657.40),
    ("Team Billing Group 6 hrs", 29905.00, 2080.00),
    ("Team Group 16 hrs", 19306.19, 618.20),
    ("Team Group 6 hrs", 8969.73, 0.00),
    ("Team Group 20 hrs", 2635.00, 0.00),
    ("Team 24-25 Billing Group 16hrs", 84.48, 0.00),
]
r = 5
for name, rev, disc in TEAM:
    lc = td.cell(row=r, column=2, value=name)
    lc.fill = lab; lc.font = Font(name=F, size=10, color="1F2937"); lc.border = box
    lc.alignment = Alignment(horizontal="left", indent=1)
    for col, val in ((3, rev), (4, disc)):
        c = td.cell(row=r, column=col, value=val)
        c.fill = white; c.font = inp; c.number_format = MONEY2; c.border = box
        c.alignment = Alignment(horizontal="center")
    r += 1

tr = r
lc = td.cell(row=tr, column=2, value="Team tuition subtotal")
lc.fill = purp; lc.font = hdrf; lc.border = box
lc.alignment = Alignment(horizontal="left", indent=1)
for col in (3, 4):
    letter = td.cell(row=1, column=col).column_letter
    c = td.cell(row=tr, column=col, value=f"=SUM({letter}5:{letter}{tr-1})")
    c.fill = totf; c.font = tot; c.number_format = MONEY2; c.border = box
    c.alignment = Alignment(horizontal="center")

r = tr + 2
for label, val in (("Cat1 Team total (from report)", 517731.45),
                   ("Less team tuition", None),
                   ("Meet entry fees & other (residual)", None)):
    lc = td.cell(row=r, column=2, value=label)
    lc.font = wbf
    lc.alignment = Alignment(horizontal="right")
    if val is not None:
        c = td.cell(row=r, column=3, value=val)
    elif label.startswith("Less"):
        c = td.cell(row=r, column=3, value=f"=-C{tr}")
    else:
        c = td.cell(row=r, column=3, value=f"=C{r-2}+C{r-1}")
    c.fill = white; c.font = tot; c.number_format = MONEY2; c.border = box
    c.alignment = Alignment(horizontal="center")
    r += 1

r += 1
for text, font in [
    ("What sits in the residual", Font(name=F, size=11, bold=True, color=WHITE)),
    ("Meet entry fees collected on behalf of host gyms — Interstate Flipin $12,280 · Commonwealth Cup 2024 $10,562 ·", w),
    ("All American Challenge 2024 $9,969 · NGUNQ $9,940 · Palmetto $7,050 · Rhode Island $14,468 across editions ·", w),
    ("Maine State Clinic $4,060 · Mens Regionals $2,101, plus roughly twenty smaller meets.", w),
    ("American Flyers Cup 2023 $43.75 and 2024 $9,401.53 — the gym's own hosted meet, on the gym's books.", w),
    ("An unnamed activity line of $30,774.29 that needs identifying.", w),
    ("", w),
    ("Treat meet fees as pass-through: money in, money out to host gyms. Near-zero gross margin. If they stay", Font(name=F, size=10, bold=True, color=AMBER)),
    ("blended into team revenue, team gross margin will read far better than it is.", w),
] :
    td.cell(row=r, column=2, value=text).font = font
    r += 1

for sheet in wbk.worksheets:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

wbk.save(OUT)
tot_rev = sum(d[2] for d in DATA)
tot_disc = sum(d[3] for d in DATA)
print("wrote", OUT)
print("sum rows      = %.2f" % tot_rev)
print("report total  = %.2f" % GRAND)
print("variance      = %.2f" % (tot_rev - GRAND))
print("discount rows = %.2f  report = %.2f  var = %.2f" % (tot_disc, GRAND_DISC, tot_disc - GRAND_DISC))
