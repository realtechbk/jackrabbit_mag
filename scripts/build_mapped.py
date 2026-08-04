import sys
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

SRC, OUT = sys.argv[1], sys.argv[2]

DARK, PURPLE, LABELBG, CALCBG, WHITE = "2F3641", "8B3DD7", "D6DEF5", "DCE3F2", "FFFFFF"
AMBER, RED = "FFD479", "FF8A8A"
F = "Arial"

w    = Font(name=F, size=10, color=WHITE)
wbf  = Font(name=F, size=10, bold=True, color=WHITE)
hdrf = Font(name=F, size=10, bold=True, color=WHITE)
inp  = Font(name=F, size=9, color="0000FF")
cal  = Font(name=F, size=9, color="000000")
tot  = Font(name=F, size=10, bold=True, color="000000")

dark  = PatternFill("solid", fgColor=DARK)
purp  = PatternFill("solid", fgColor=PURPLE)
lab   = PatternFill("solid", fgColor=LABELBG)
calc  = PatternFill("solid", fgColor=CALCBG)
white = PatternFill("solid", fgColor=WHITE)
totf  = PatternFill("solid", fgColor="B9C4DC")

thin = Side(style="thin", color="9AA3B2")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

MONEY = '$#,##0;($#,##0);-'
NUM   = '#,##0;(#,##0);-'
PCT   = '0.0%;(0.0%);-'

# ---------------- load + map ----------------
df = pd.read_excel(SRC, sheet_name=0)
df.columns = [str(c).replace("\n", " ").strip() for c in df.columns]

df["Cat1"] = df["Cat1"].fillna("").astype(str).str.strip()
df["Cat2"] = df["Cat2"].fillna("").astype(str).str.strip()
df["Cat3"] = df["Cat3"].fillna("").astype(str).str.strip()
df["Tuition"] = pd.to_numeric(df["Tuition"], errors="coerce").fillna(0.0)
for c in ("Size", "Max"):
    df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0).astype(int)

KEYMAP = {
    ("Recreational", "Pre-School"):        "Recreational -- Preschool (The Jungle)",
    ("Recreational", "Parent-Child"):      "Recreational -- Preschool (The Jungle)",
    ("Recreational", "Rec Girls"):         "Recreational -- Girls Wings",
    ("Recreational", "Rec Boys"):          "Recreational -- Boys Wings",
    ("Recreational", "Tumble (Combined)"): "Recreational -- Tumbling",
    ("Team", ""):                          "Competitive -- American Flyers Teams",
    ("Staff", ""):                         "EXCLUDE -- Staff placeholder",
}
CAT3_OVERRIDE = {
    "Girls PT1": "Competitive -- Pre-Team",
    "Boys PT2":  "Competitive -- Pre-Team",
}

SERVICE_ROWS = [
    "Recreational -- Preschool (The Jungle)",
    "Recreational -- Girls Wings",
    "Recreational -- Boys Wings",
    "Recreational -- Tumbling",
    "Competitive -- Pre-Team",
    "Competitive -- American Flyers Teams",
    "EXCLUDE -- Staff placeholder",
]


def map_row(r):
    if r["Cat3"] in CAT3_OVERRIDE:
        return CAT3_OVERRIDE[r["Cat3"]]
    return KEYMAP.get((r["Cat1"], r["Cat2"]), "UNMAPPED")


df["MapsTo"] = df.apply(map_row, axis=1)
df["CatKey"] = df["Cat1"] + " | " + df["Cat2"].replace("", "(blank)")
df["Summer"] = df["Session"].fillna("").astype(str).str.contains("Summer")
df["Billable"] = ((df["Tuition"] > 0) & (~df["MapsTo"].str.startswith("EXCLUDE"))).map({True: "Yes", False: "No"})


def flag(r):
    if r["MapsTo"] == "UNMAPPED":
        return "UNMAPPED"
    if r["MapsTo"].startswith("EXCLUDE"):
        return "EXCLUDED - not revenue"
    if r["Tuition"] == 0 and r["Cat1"] == "Team":
        return "TEAM PRACTICE GRP - $0"
    if r["Tuition"] == 0:
        return "NO TUITION"
    if r["Summer"]:
        return "OK - summer session"
    return "OK"


df["Flag"] = df.apply(flag, axis=1)
df = df.sort_values(["MapsTo", "Cat3", "Class"], kind="stable").reset_index(drop=True)

wbk = Workbook()


def canvas(ws, nr, nc):
    ws.sheet_view.showGridLines = False
    for r in range(1, nr + 1):
        for c in range(1, nc + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = dark
            cell.font = w


def H(ws, row, col, text, width=None):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = purp
    c.font = hdrf
    c.border = box
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    if width:
        ws.column_dimensions[c.column_letter].width = width
    return c


# ============================ Findings ============================
fs = wbk.active
fs.title = "Findings"
canvas(fs, 78, 8)
fs.column_dimensions['A'].width = 2
fs.column_dimensions['B'].width = 4
fs.column_dimensions['C'].width = 112

fs['B1'] = "Jackrabbit Class List — Reconciliation Findings"
fs['B1'].font = Font(name=F, size=16, bold=True, color=WHITE)
fs['B2'] = f"Source: ClassesListReport.xls · {len(df)} classes · exported 28 July 2026"
fs['B2'].font = Font(name=F, size=10, italic=True, color="C9D1E0")

blocks = [
    ("HEADLINE", AMBER, [
        "The categories DO support your nine service rows — with two structural exceptions you need to decide on.",
        "But this export contains NO 2023, 2024 or 2025 classes. It is the current session only.",
    ]),
    ("1. Category 1 separates Rec from Team cleanly", None, [
        "Recreational 155  ·  Team 18  ·  Staff 3.  My earlier read off your screenshot was wrong — Cat1 is not",
        "uniformly 'Recreational'. It does carry the top-level split.",
    ]),
    ("2. Category 2 is the service-level key", None, [
        "Rec Girls 86  ·  Pre-School 36  ·  Rec Boys 12  ·  Tumble (Combined) 12  ·  Parent-Child 9  ·  blank 21.",
        "These map onto four of your rows with no ambiguity. Parent-Child (First Flight) folds into Preschool.",
        "The 21 blanks are exactly the 18 Team + 3 Staff classes — Team carries no Cat2 or Cat3 at all.",
    ]),
    ("3. TEAM IS DOUBLE-ENROLLED BY DESIGN — the important one", AMBER, [
        "Only 7 of the 18 Team classes carry tuition. They are hour-tier billing groups:",
        "     1 Day $255  ·  6 hrs $310  ·  2 days $350  ·  9 hrs $385  ·  12 hr $440  ·  16 hrs $485  ·  20 hrs $543",
        "The other 11 are $0 containers — 'MAG Team 2026-27' (77 enrolled) plus practice groups (Xcel Silver/Gold/",
        "Platinum/Diamond, WDP L3/L4/Optionals, Mens Devo, Womens Devo).",
        "Team athletes sit in a billing group AND a practice group AND the roster. Summing enrollment across all 18",
        "would triple-count them. Team revenue = the 7 billing groups only (74 enrolled). Use practice groups for",
        "headcount and level mix, never for revenue.",
        "Consequence: you CANNOT split team revenue girls vs boys. Billing groups are gender-blind. Headcount can be",
        "inferred (Mens Devo 13 vs everything else) but the dollars cannot.",
    ]),
    ("4. Pre-Team barely exists as a service line", None, [
        "Only 2 classes are tagged PT in Cat3: 'Hot Shots' (Girls PT1, 9 enrolled, $310) and 'Boys Advanced 2 hr'",
        "(Boys PT2, 1 enrolled, $200). Both sit under Rec Girls / Rec Boys in Cat2, so I applied a Cat3 override.",
        "10 students total. Consider folding this row into the rec rows — it will not carry a meaningful trend.",
    ]),
    ("5. Summer Intensives are a separate product priced differently", None, [
        "11 classes in '2026 Summer Intensive' — the website's Summer Mini-Sessions. Flat 6-week fees ($183–$300),",
        "not monthly tuition. They carry the same Cat2 values as the year-round classes, so they will blend into",
        "your rec rows and distort Avg. Transaction Value unless separated. Flagged 'OK - summer session'.",
    ]),
    ("6. Thirty-six classes carry $0 tuition", None, [
        "22 Recreational  ·  11 Team  ·  3 Staff. Waitlists, roster containers, practice groups and staff",
        "placeholders. Excluded from pricing via the 'Billable' column so they do not drag the averages to zero.",
    ]),
    ("7. NO HISTORICAL DATA IN THIS EXPORT", RED, [
        "Session mix: 2026-27 Rec 144  ·  2026-27 Team 18  ·  2026 Summer Intensive 11  ·  2025-26 Rec 1  ·  blank 2.",
        "There is not a single 2023, 2024 or 2025 class here. This file gives you current price and current",
        "enrollment — nothing for the three-year comparison.",
        "The 2023–2025 numbers must come from the revenue and transaction reports filtered by date range, not from",
        "the class list. Re-run the class list with archived sessions included if you also want historical prices.",
    ]),
    ("8. Fill rate — a genuine finding, available right now", None, [
        "Tumbling 81.6%  ·  Preschool 77.0%  ·  Rec Girls 69.8%  ·  Parent-Child 58.3%  ·  Rec Boys 51.5%",
        "Rec Boys is running about half empty across 12 classes. Since coach cost is fixed once a class runs, those",
        "empty seats are pure lost margin — likely the single highest-leverage item in the whole analysis.",
    ]),
    ("WHAT I NEED FROM YOU NEXT", AMBER, [
        "a)  Decide: keep Pre-Team as its own row, or fold into rec?",
        "b)  Decide: Summer Intensives as their own row, or blended into the rec rows?",
        "c)  Pull the revenue reports for calendar 2023, 2024 and 2025 (Revenue Summary or Class/Event Revenue),",
        "     on a consistent basis (billed or collected), so I can fill the three-year sheet.",
    ]),
]

r = 4
for title, color, lines in blocks:
    c = fs.cell(row=r, column=2, value=title)
    c.font = Font(name=F, size=11, bold=True, color=color or WHITE)
    r += 1
    for ln in lines:
        fs.cell(row=r, column=3, value=ln).font = w
        r += 1
    r += 1

# ============================ Class Export ============================
ce = wbk.create_sheet("Class Export (mapped)")
ncols = 18
canvas(ce, len(df) + 12, ncols + 2)
ce.column_dimensions['A'].width = 2

ce['B1'] = "Class Export — mapped to service rows"
ce['B1'].font = Font(name=F, size=13, bold=True, color=WHITE)

FR = 4
cols = [
    ("Class", "Class", 40),
    ("Cat1", "Cat1", 13),
    ("Cat2", "Cat2", 17),
    ("Cat3", "Cat3", 11),
    ("Session", "Session", 20),
    ("Status", "Status", 9),
    ("Instructors", "Instructors", 24),
    ("Days", "Days", 7),
    ("Start Time", "Start Time", 10),
    ("Duration", "Duration", 9),
    ("Tuition", "Tuition", 10),
    ("Enrolled", "Size", 9),
    ("Max", "Max", 8),
    ("Cat Key", "CatKey", 26),
    ("Maps To", "MapsTo", 44),
    ("Billable", "Billable", 9),
    ("Flag", "Flag", 24),
]
for i, (label, _, width) in enumerate(cols):
    H(ce, FR, 2 + i, label, width)

for j, (_, src, _) in enumerate(cols):
    colno = 2 + j
    for i, val in enumerate(df[src].tolist()):
        cell = ce.cell(row=FR + 1 + i, column=colno)
        if src in ("Start Time",):
            cell.value = str(val) if pd.notna(val) else ""
        else:
            cell.value = val
        cell.border = box
        if src in ("CatKey", "MapsTo", "Billable", "Flag"):
            cell.fill = calc
            cell.font = cal
        else:
            cell.fill = white
            cell.font = inp
        if src == "Tuition":
            cell.number_format = MONEY
            cell.alignment = Alignment(horizontal="center")
        elif src in ("Size", "Max"):
            cell.number_format = NUM
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.alignment = Alignment(horizontal="left", indent=1)

LR = FR + len(df)
ce.freeze_panes = "C5"

# ============================ Service Summary ============================
ss = wbk.create_sheet("Service Summary")
canvas(ss, 40, 12)
ss.column_dimensions['A'].width = 2
ss['B1'] = "Service Summary — live from the Class Export tab"
ss['B1'].font = Font(name=F, size=13, bold=True, color=WHITE)
ss['B2'] = "Billable classes only ($0 and staff placeholders excluded from price columns)"
ss['B2'].font = Font(name=F, size=10, italic=True, color="C9D1E0")

heads = [("Service Row", 44), ("Classes", 10), ("Billable classes", 14), ("Enrolled", 10),
         ("Capacity", 10), ("Fill %", 10), ("Min monthly", 12), ("Avg monthly", 12), ("Max monthly", 12)]
for i, (label, width) in enumerate(heads):
    H(ss, 4, 2 + i, label, width)

EX = "'Class Export (mapped)'"
M = f"{EX}!$P${FR+1}:$P${LR}"      # Maps To
B = f"{EX}!$Q${FR+1}:$Q${LR}"      # Billable
T = f"{EX}!$L${FR+1}:$L${LR}"      # Tuition
S = f"{EX}!$M${FR+1}:$M${LR}"      # Size
X = f"{EX}!$N${FR+1}:$N${LR}"      # Max

r = 5
for name in SERVICE_ROWS:
    c = ss.cell(row=r, column=2, value=name)
    c.fill = lab
    c.font = Font(name=F, size=10, bold=True, color="1F2937")
    c.border = box
    c.alignment = Alignment(horizontal="left", indent=1)

    formulas = [
        (3,  f'=COUNTIF({M},$B{r})', NUM),
        (4,  f'=COUNTIFS({M},$B{r},{B},"Yes")', NUM),
        (5,  f'=SUMIF({M},$B{r},{S})', NUM),
        (6,  f'=SUMIF({M},$B{r},{X})', NUM),
        (7,  f'=IFERROR(E{r}/F{r},0)', PCT),
        (8,  f'=IFERROR(_xlfn.MINIFS({T},{M},$B{r},{B},"Yes"),0)', MONEY),
        (9,  f'=IFERROR(AVERAGEIFS({T},{M},$B{r},{B},"Yes"),0)', MONEY),
        (10, f'=IFERROR(_xlfn.MAXIFS({T},{M},$B{r},{B},"Yes"),0)', MONEY),
    ]
    for col, formula, fmt in formulas:
        cc = ss.cell(row=r, column=col, value=formula)
        cc.fill = calc
        cc.font = cal
        cc.number_format = fmt
        cc.border = box
        cc.alignment = Alignment(horizontal="center")
    r += 1

tl = ss.cell(row=r, column=2, value="TOTAL (excl. staff placeholders)")
tl.fill = purp
tl.font = hdrf
tl.border = box
tl.alignment = Alignment(horizontal="left", indent=1)
for col, formula, fmt in [
    (3,  f'=SUM(C5:C{r-2})', NUM),
    (4,  f'=SUM(D5:D{r-2})', NUM),
    (5,  f'=SUM(E5:E{r-2})', NUM),
    (6,  f'=SUM(F5:F{r-2})', NUM),
    (7,  f'=IFERROR(E{r}/F{r},0)', PCT),
]:
    cc = ss.cell(row=r, column=col, value=formula)
    cc.fill = totf
    cc.font = tot
    cc.number_format = fmt
    cc.border = box
    cc.alignment = Alignment(horizontal="center")

r += 2
notes = [
    ("Reading these numbers", Font(name=F, size=11, bold=True, color=WHITE)),
    ("Team enrolled counts the 7 billing groups plus the 11 $0 practice/roster groups, so it overstates unique athletes.", w),
    ("Unique team athletes = the 'MAG Team 2026-27' roster (77). Billing-group enrolment totals 74.", w),
    ("Tuition is monthly except the 11 Summer Intensive classes, which are flat 6-week fees.", w),
    ("Source: Jackrabbit Class List export, 176 classes, 28 July 2026. Mapping rules on the Crosswalk tab.", Font(name=F, size=9, italic=True, color="C9D1E0")),
]
for text, font in notes:
    ss.cell(row=r, column=2, value=text).font = font
    r += 1

# ============================ Crosswalk ============================
cw = wbk.create_sheet("Crosswalk")
canvas(cw, 40, 8)
cw.column_dimensions['A'].width = 2
cw['B1'] = "Mapping rules applied"
cw['B1'].font = Font(name=F, size=13, bold=True, color=WHITE)

H(cw, 4, 2, "Cat1", 16)
H(cw, 4, 3, "Cat2", 20)
H(cw, 4, 4, "Cat3 override", 16)
H(cw, 4, 5, "Maps to service row", 44)
H(cw, 4, 6, "Classes", 10)

rules = [
    ("Recreational", "Pre-School", "", "Recreational -- Preschool (The Jungle)", 36),
    ("Recreational", "Parent-Child", "", "Recreational -- Preschool (The Jungle)", 9),
    ("Recreational", "Rec Girls", "", "Recreational -- Girls Wings", 85),
    ("Recreational", "Rec Boys", "", "Recreational -- Boys Wings", 11),
    ("Recreational", "Tumble (Combined)", "", "Recreational -- Tumbling", 12),
    ("Recreational", "Rec Girls", "Girls PT1", "Competitive -- Pre-Team", 1),
    ("Recreational", "Rec Boys", "Boys PT2", "Competitive -- Pre-Team", 1),
    ("Team", "(blank)", "", "Competitive -- American Flyers Teams", 18),
    ("Staff", "(blank)", "", "EXCLUDE -- Staff placeholder", 3),
]
r = 5
for a, b, c_, d, n in rules:
    for col, val in ((2, a), (3, b), (4, c_ or "—"), (5, d), (6, n)):
        cc = cw.cell(row=r, column=col, value=val)
        cc.fill = white if col < 5 else calc
        cc.font = Font(name=F, size=10, color="1F2937")
        cc.border = box
        cc.alignment = Alignment(horizontal="center" if col == 6 else "left", indent=1)
    r += 1

r += 1
cw.cell(row=r, column=2, value="Rows not yet in your nine-row structure:").font = Font(name=F, size=10, bold=True, color=AMBER)
r += 1
for text in [
    "Ancillary -- Open Gym  ·  Ancillary -- Annual Membership Fees  ·  Events -- American Flyers Cup",
    "None of these are classes, so they do not appear in this export. They live in the transaction sub-types instead.",
]:
    cw.cell(row=r, column=2, value=text).font = w
    r += 1

for ws_ in wbk.worksheets:
    ws_.page_setup.orientation = "landscape"
    ws_.page_setup.fitToWidth = 1
    ws_.page_setup.fitToHeight = 0
    ws_.sheet_properties.pageSetUpPr.fitToPage = True

wbk.save(OUT)
print("wrote", OUT, "rows:", len(df))
print(df["MapsTo"].value_counts().to_string())
print()
print(df["Flag"].value_counts().to_string())
