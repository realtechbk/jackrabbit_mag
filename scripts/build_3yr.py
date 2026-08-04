import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = sys.argv[1]

DARK, PURPLE, LABELBG, CALCBG, WHITE = "2F3641", "8B3DD7", "D6DEF5", "DCE3F2", "FFFFFF"
AMBER, REDT = "FFD479", "FF9B9B"
F = "Arial"

w    = Font(name=F, size=10, color=WHITE)
wbf  = Font(name=F, size=10, bold=True, color=WHITE)
hdrf = Font(name=F, size=10, bold=True, color=WHITE)
inp  = Font(name=F, size=10, color="0000FF")
cal  = Font(name=F, size=10, color="000000")
tot  = Font(name=F, size=10, bold=True, color="000000")

dark  = PatternFill("solid", fgColor=DARK)
purp  = PatternFill("solid", fgColor=PURPLE)
lab   = PatternFill("solid", fgColor=LABELBG)
calc  = PatternFill("solid", fgColor=CALCBG)
white = PatternFill("solid", fgColor=WHITE)
totf  = PatternFill("solid", fgColor="B9C4DC")
flagf = PatternFill("solid", fgColor="E8D9F5")
exclf = PatternFill("solid", fgColor="D9D9D9")

thin = Side(style="thin", color="9AA3B2")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

M2  = '$#,##0.00;($#,##0.00);-'
M0  = '$#,##0;($#,##0);-'
PCT = '0.0%;(0.0%);-'
GRW = '+0.0%;-0.0%;0.0%'
NUM1 = '#,##0;(#,##0);-'

# ---- verified figures, all reconciled to the Jackrabbit Revenue Summary totals ----
# label, group, 2023, 2024, 2025, 2026 YTD, note
ROWS = [
    ("Recreational -- Girls Wings", "R", 748491.02, 999839.46, 975278.37, 505818.42,
     "Cat2 Rec Girls. Includes pre-team girls from 2024 (see drift tab)."),
    ("Recreational -- Preschool (The Jungle)", "R", 343335.26, 381504.55, 367014.79, 194726.85,
     "Cat2 Pre-School + Parent-Child."),
    ("Recreational -- Boys Wings", "R", 86323.49, 84553.60, 85489.15, 40964.00,
     "Cat2 Rec Boys. Flat for three straight years."),
    ("Recreational -- Tumbling", "R", 83307.09, 102217.79, 89018.18, 36851.55,
     "Cat2 Tumble (Combined)."),
    ("Recreational -- Pre-Team", "D", 26248.25, 0.00, 0.00, 0.00,
     "Cat2 'Pre-Team (Invite Only)' existed in 2023 ONLY. Folded into Wings rows after."),
    ("Recreational -- Unclassified (no Cat2)", "F", 21744.45, 27885.43, 19636.96, 9468.75,
     "Cat1 Recreational with blank Cat2. Cannot be attributed to a service."),
    ("Competitive -- Team (tuition + meet fees)", "R", 517731.45, 552485.77, 564294.62, 259212.55,
     "Cat1 Team. 2023 splits 377,627 tuition / 140,104 meet fees; later years not yet split."),
    ("Ancillary -- Annual Membership Fees", "R", 58293.50, 61640.05, 52776.95, 21575.30,
     "Cat1 Annual Membership. At $45/student this is a proxy for active student count."),
    ("Ancillary -- Pro Shop", "R", 28031.65, 28789.86, 34975.39, 15754.09,
     "Cat1 ProShop. Growing steadily."),
    ("Ancillary -- Open Gym", "R", 30059.50, 23357.15, 28128.56, 20980.25,
     "Cat1 Open Gym. 2026 already at 75% of full-year 2025 by July."),
    ("Ancillary -- Birthday Parties", "D", 45767.05, 19038.20, 0.00, 0.00,
     "Cat1 Birthday Parties. Wound down to zero — appears discontinued."),
    ("Ancillary -- Private Lessons", "R", 3815.71, 5645.13, 8318.25, 6055.90,
     "Cat1 Private Lessons. More than doubled 2023-25, still climbing."),
    ("Ancillary -- Gift Certificates", "D", 310.00, 270.00, 0.00, 0.00,
     "Cat1 Gift Certificate. Gone from 2025."),
    ("Ancillary -- Non-Instructional Time", "D", 0.00, 0.00, 340.00, 0.00,
     "Cat1 present in 2025 only."),
    ("Other -- Tumble (stray Cat1)", "F", 0.00, 0.00, 17.80, 0.00,
     "Stray Cat1 'Tumble' in 2025 — a mis-tag, belongs under Recreational."),
    ("EXCLUDE -- Staff", "X", 2660.10, 10226.82, 3824.77, 1998.03,
     "Cat1 Staff. Not customer revenue."),
    ("EXCLUDE -- Unapplied Payments", "X", 172.00, -100.00, 11920.58, 12817.63,
     "Cat1 --Unapplied Payments--. $12,818 unapplied by July 2026 and still growing."),
]

REPORTED = {"2023": 1996290.52, "2024": 2297353.81, "2025": 2241034.37, "2026": 1126223.32}
DISCOUNTS = {"2023": 128402.40, "2024": 145130.58, "2025": 148452.08, "2026": 80381.12}
YEARS = ["2023", "2024", "2025", "2026"]
YTD_FRACTION = 209 / 365.0    # 1 Jan to 28 Jul 2026

wbk = Workbook()


def canvas(ws, nr, nc):
    ws.sheet_view.showGridLines = False
    for r in range(1, nr + 1):
        for c in range(1, nc + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = dark
            cell.font = w


def H(ws, row, col, text, width=None, align="center"):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = purp
    c.font = hdrf
    c.border = box
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True,
                            indent=1 if align == "left" else 0)
    if width:
        ws.column_dimensions[c.column_letter].width = width
    return c


# ==================== 3-Year Comparison ====================
ws = wbk.active
ws.title = "3-Year by Service"
canvas(ws, 62, 15)
ws.column_dimensions['A'].width = 2

ws['B1'] = "Maine Academy of Gymnastics — Revenue by Service Line, 2023-2026 YTD"
ws['B1'].font = Font(name=F, size=15, bold=True, color=WHITE)
ws['B2'] = "Source: Jackrabbit Revenue Summary, one report per period, all run 28 July 2026. Every period reconciles to its stated total with zero variance."
ws['B2'].font = Font(name=F, size=10, italic=True, color="C9D1E0")
ws['B3'] = "2026 covers 1 Jan to 28 Jul only (209 of 365 days, 57.3%). It is NOT comparable to the full years without adjustment — see the pace column."
ws['B3'].font = Font(name=F, size=10, bold=True, color=AMBER)

al = ws.cell(row=2, column=12, value="YTD fraction of year:")
al.font = wbf
al.alignment = Alignment(horizontal="right")
av = ws.cell(row=2, column=13, value=YTD_FRACTION)
av.fill = PatternFill("solid", fgColor="FFFF00")
av.font = Font(name=F, size=10, bold=True, color="000000")
av.number_format = PCT
av.border = box
av.alignment = Alignment(horizontal="center")
ws.cell(row=3, column=12, value="209 of 365 days. Straight-line — ignores seasonality.").font = \
    Font(name=F, size=8, italic=True, color="C9D1E0")
ws.column_dimensions['L'].width = 24
ws.column_dimensions['M'].width = 10

H(ws, 4, 2, "Service Line", 42, "left")
H(ws, 4, 3, "2023", 13)
H(ws, 4, 4, "2024", 13)
H(ws, 4, 5, "2025", 13)
H(ws, 4, 6, "2026 YTD\n(to 28 Jul)", 13)
H(ws, 4, 7, "24 vs 23", 10)
H(ws, 4, 8, "25 vs 24", 10)
H(ws, 4, 9, "26 pace\nvs 2025", 11)
H(ws, 4, 10, "% of 2025", 10)
H(ws, 4, 11, "Note", 58, "left")

FR = 5
r = FR
for label, grp, v23, v24, v25, v26, note in ROWS:
    fill = {"R": lab, "D": flagf, "F": flagf, "X": exclf}[grp]
    lc = ws.cell(row=r, column=2, value=label)
    lc.fill = fill
    lc.font = Font(name=F, size=10, bold=True, color="1F2937")
    lc.border = box
    lc.alignment = Alignment(horizontal="left", indent=1)

    for col, val in ((3, v23), (4, v24), (5, v25), (6, v26)):
        c = ws.cell(row=r, column=col, value=val)
        c.fill = white
        c.font = inp
        c.number_format = M2
        c.border = box
        c.alignment = Alignment(horizontal="center")

    for col, formula in ((7, f'=IF(OR(C{r}=0,D{r}=0),"n/a",D{r}/C{r}-1)'),
                         (8, f'=IF(OR(D{r}=0,E{r}=0),"n/a",E{r}/D{r}-1)'),
                         (9, f'=IF(OR(E{r}=0,F{r}=0),"n/a",(F{r}/$M$2)/E{r}-1)')):
        c = ws.cell(row=r, column=col, value=formula)
        c.fill = calc
        c.font = cal
        c.number_format = GRW
        c.border = box
        c.alignment = Alignment(horizontal="center")

    c = ws.cell(row=r, column=10, value=f"=IFERROR(E{r}/$E${FR + len(ROWS)},0)")
    c.fill = calc
    c.font = cal
    c.number_format = PCT
    c.border = box
    c.alignment = Alignment(horizontal="center")

    c = ws.cell(row=r, column=11, value=note)
    c.fill = calc
    c.font = Font(name=F, size=9, color="1F2937")
    c.border = box
    c.alignment = Alignment(horizontal="left", indent=1)
    r += 1

LR = r - 1
TOTR = r
tl = ws.cell(row=TOTR, column=2, value="TOTAL REVENUE")
tl.fill = purp
tl.font = hdrf
tl.border = box
tl.alignment = Alignment(horizontal="left", indent=1)
for col in (3, 4, 5, 6):
    letter = ws.cell(row=1, column=col).column_letter
    c = ws.cell(row=TOTR, column=col, value=f"=SUM({letter}{FR}:{letter}{LR})")
    c.fill = totf
    c.font = tot
    c.number_format = M2
    c.border = box
    c.alignment = Alignment(horizontal="center")
for col, formula in ((7, f"=D{TOTR}/C{TOTR}-1"),
                     (8, f"=E{TOTR}/D{TOTR}-1"),
                     (9, f"=(F{TOTR}/$M$2)/E{TOTR}-1")):
    c = ws.cell(row=TOTR, column=col, value=formula)
    c.fill = totf
    c.font = tot
    c.number_format = GRW
    c.border = box
    c.alignment = Alignment(horizontal="center")
c = ws.cell(row=TOTR, column=10, value=f"=IFERROR(E{TOTR}/E{TOTR},0)")
c.fill = totf; c.font = tot; c.number_format = PCT; c.border = box
c.alignment = Alignment(horizontal="center")

# annualised 2026
ar = TOTR + 1
lc = ws.cell(row=ar, column=2, value="2026 annualised (straight-line)")
lc.fill = flagf
lc.font = Font(name=F, size=10, bold=True, color="1F2937")
lc.border = box
lc.alignment = Alignment(horizontal="left", indent=1)
c = ws.cell(row=ar, column=6, value=f"=F{TOTR}/$M$2")
c.fill = PatternFill("solid", fgColor="FFF2CC")
c.font = tot; c.number_format = M2; c.border = box
c.alignment = Alignment(horizontal="center")
ws.cell(row=ar, column=11, value="Seasonality not modelled — treat as a rough pace indicator only.").font = \
    Font(name=F, size=9, italic=True, color="C9D1E0")

# reconciliation
COLS = ((3, "2023"), (4, "2024"), (5, "2025"), (6, "2026"))
r = ar + 2
ws.cell(row=r, column=2, value="RECONCILIATION").font = Font(name=F, size=11, bold=True, color=WHITE)
r += 1
rep_row = r
lc = ws.cell(row=r, column=2, value="Report stated total")
lc.font = wbf
lc.alignment = Alignment(horizontal="right")
for col, y in COLS:
    c = ws.cell(row=r, column=col, value=REPORTED[y])
    c.fill = white; c.font = tot; c.number_format = M2; c.border = box
    c.alignment = Alignment(horizontal="center")
r += 1
lc = ws.cell(row=r, column=2, value="Variance (must be zero)")
lc.font = wbf
lc.alignment = Alignment(horizontal="right")
for col, _ in COLS:
    letter = ws.cell(row=1, column=col).column_letter
    c = ws.cell(row=r, column=col, value=f"={letter}{TOTR}-{letter}{rep_row}")
    c.fill = white; c.font = tot; c.number_format = M2; c.border = box
    c.alignment = Alignment(horizontal="center")
r += 1
lc = ws.cell(row=r, column=2, value="Discounts given (from report)")
lc.font = wbf
lc.alignment = Alignment(horizontal="right")
for col, y in COLS:
    c = ws.cell(row=r, column=col, value=DISCOUNTS[y])
    c.fill = white; c.font = tot; c.number_format = M2; c.border = box
    c.alignment = Alignment(horizontal="center")
disc_row = r
r += 1
lc = ws.cell(row=r, column=2, value="Discount rate on gross")
lc.font = wbf
lc.alignment = Alignment(horizontal="right")
for col, _ in COLS:
    letter = ws.cell(row=1, column=col).column_letter
    c = ws.cell(row=r, column=col,
                value=f"={letter}{disc_row}/({letter}{TOTR}+{letter}{disc_row})")
    c.fill = white; c.font = tot; c.number_format = PCT; c.border = box
    c.alignment = Alignment(horizontal="center")
r += 1

# student proxy
r += 1
ws.cell(row=r, column=2, value="ACTIVE STUDENT PROXY (membership fees / $45)").font = Font(name=F, size=11, bold=True, color=AMBER)
r += 1
mem_row = FR + [x[0] for x in ROWS].index("Ancillary -- Annual Membership Fees")
lc = ws.cell(row=r, column=2, value="Implied active students")
lc.font = wbf
lc.alignment = Alignment(horizontal="right")
for col, _ in COLS:
    letter = ws.cell(row=1, column=col).column_letter
    c = ws.cell(row=r, column=col, value=f"={letter}{mem_row}/45")
    c.fill = white; c.font = tot; c.number_format = NUM1; c.border = box
    c.alignment = Alignment(horizontal="center")
stu_row = r
ws.cell(row=r, column=12, value="2026 is part-year — not comparable.").font = \
    Font(name=F, size=8, italic=True, color=AMBER)
r += 1
lc = ws.cell(row=r, column=2, value="Revenue per active student")
lc.font = wbf
lc.alignment = Alignment(horizontal="right")
for col, y in COLS:
    letter = ws.cell(row=1, column=col).column_letter
    c = ws.cell(row=r, column=col, value=f"=IFERROR({letter}{TOTR}/{letter}{stu_row},0)")
    c.fill = white; c.font = tot; c.number_format = M0; c.border = box
    c.alignment = Alignment(horizontal="center")
r += 1

r += 1
key = [
    ("KEY LEGEND", Font(name=F, size=10, bold=True, color=WHITE)),
    ("Blue cells are hardcoded figures taken directly from the Jackrabbit reports. Growth and share columns calculate.", w),
    ("Purple-shaded rows are discontinued lines or categorisation artefacts, not clean comparatives.", w),
    ("Grey rows are excluded from any operating view.", w),
]
for text, font in key:
    ws.cell(row=r, column=2, value=text).font = font
    r += 1

ws.freeze_panes = "C5"

# ==================== Findings ====================
fs = wbk.create_sheet("Findings")
canvas(fs, 76, 6)
fs.column_dimensions['A'].width = 2
fs.column_dimensions['B'].width = 4
fs.column_dimensions['C'].width = 112

fs['B1'] = "Three-Year Findings"
fs['B1'].font = Font(name=F, size=16, bold=True, color=WHITE)
fs['B2'] = "2023: $1,996,291  ·  2024: $2,297,354 (+15.1%)  ·  2025: $2,241,034 (-2.5%)"
fs['B2'].font = Font(name=F, size=11, color=AMBER)

blocks = [
    ("THE HEADLINE", AMBER, [
        "Revenue grew 15.1% in 2024, then fell 2.5% in 2025. But that small decline hides a much larger one:",
        "implied active students fell from 1,370 to 1,173 — down 14.4%.",
        "Revenue held up because price per student rose, not because the business held its customers.",
        "Revenue per active student: $1,541 (2023) -> $1,677 (2024) -> $1,911 (2025), up 24% in two years.",
    ]),
    ("1. The 2025 story is price masking volume", None, [
        "This is the single most important finding. A 2.5% revenue dip reads as a flat year; a 14.4% drop in students",
        "is a retention problem. The gym raised effective price enough to cover it, which works once and is hard to",
        "repeat. If enrolment does not recover, 2026 revenue falls unless price rises again.",
        "Caveat: the student proxy assumes the $45 fee is charged once per active student per year. Verify against an",
        "Enrollment Snapshot before presenting the figure — it is directionally sound but worth confirming.",
    ]),
    ("2. Boys Wings has been flat for three straight years", None, [
        "$86,323 -> $84,554 -> $85,489. Essentially zero movement while girls' revenue grew 30%.",
        "Combine with the 51.6% fill rate from the current class list and the picture is consistent: the boys'",
        "programme is subscale and static. Either invest in it or consolidate the classes and free the coach hours.",
    ]),
    ("3. Girls Wings drove all the growth, then stalled", None, [
        "$748,491 -> $999,839 (+33.6%) -> $975,278 (-2.5%). 43.5% of 2025 revenue.",
        "The 2024 surge is where the whole company's growth came from. The 2025 dip mirrors the group total, so",
        "whatever happened in 2025 happened here first.",
    ]),
    ("4. Birthday Parties was killed off — a $46k line", REDT, [
        "$45,767 (2023) -> $19,038 (2024) -> absent (2025). Gift Certificates went the same way ($310 -> $270 -> 0).",
        "The magnitudes rule out a rename: the only new 2025 category, 'Non-Instructional time', is $340.",
        "This was 2.3% of revenue with almost no coach cost attached. Worth asking the client whether this was a",
        "deliberate exit or drift — and note the website still shows party photographs but has no Parties page.",
    ]),
    ("5. Private Lessons is the quiet grower", None, [
        "$3,816 -> $5,645 -> $8,318. Up 118% in two years off a small base.",
        "One-to-one delivery, premium pricing, no capacity constraint beyond coach time. Given that staffing is the",
        "stated bottleneck this may not be scalable, but the demand signal is real.",
    ]),
    ("6. Team is stable and under-analysed", None, [
        "$517,731 -> $552,486 -> $564,295. Steady low single-digit growth, 25.2% of 2025 revenue.",
        "But 2023's split showed only $377,627 of that was tuition — $140,104 was meet entry fees, largely",
        "pass-through to host gyms. The 2024 and 2025 reports need the same split before team margin means anything.",
    ]),
    ("7. Category drift breaks strict comparability", AMBER, [
        "Pre-Team (Invite Only) was its own Cat2 in 2023 ($26,248) and disappeared in 2024-25, folding into the",
        "Wings rows. So Girls Wings growth from 2023 to 2024 is overstated by roughly the pre-team amount.",
        "A stray Cat1 'Tumble' ($17.80) appeared in 2025 and should sit under Recreational.",
        "Unclassified Recreational revenue with no Cat2 runs $19,637-$27,885 a year — around 1% that cannot be",
        "attributed to any service.",
    ]),
    ("8. Housekeeping worth raising", None, [
        "Unapplied payments: $172 (2023) -> -$100 (2024) -> $11,921 (2025). That is real cash sitting unallocated",
        "against families' accounts and should be cleared.",
        "Staff category: $2,660 -> $10,227 -> $3,825. The 2024 spike needs an explanation.",
        "Discounting is rising: $128,402 -> $145,131 -> $148,452, running 6.0-6.2% of gross throughout.",
    ]),
    ("THE 2026 FILE IS UNUSABLE", REDT, [
        "RPT_515608_RSR_54371_2026.pdf has a date range of 1/1/2026 to 12/31/2025 — the end date precedes the start",
        "date, so the report returned $0.00 with no rows. Re-run it as 1/1/2026 to 12/31/2026.",
        "Note that 2026 will be a part-year figure and not comparable to the full years without annualising.",
    ]),
    ("WHAT IS STILL MISSING", AMBER, [
        "a)  Transaction counts. Every Revenue Summary shows Enrollment as 0, so '# Sales Transactions' and",
        "     'Avg. Sales Transaction Value' cannot be filled from these reports. Pull Transaction Listing or",
        "     Transaction Summary for the same three date ranges.",
        "b)  Cost of sales. Coach hourly rates plus class hours, from payroll. Nothing in Jackrabbit provides it.",
        "c)  The team tuition versus meet-fee split for 2024 and 2025.",
    ]),
]

r = 4
for title, color, lines in blocks:
    c = fs.cell(row=r, column=2, value=title)
    c.font = Font(name=F, size=11, bold=True, color=color or WHITE)
    r += 1
    for ln in lines:
        fs.cell(row=r, column=3, value=ln).font = w
        r += 1
    r += 1

# ==================== Category Drift ====================
cd = wbk.create_sheet("Category Drift")
canvas(cd, 40, 7)
cd.column_dimensions['A'].width = 2
cd['B1'] = "Cat1 categories present by year"
cd['B1'].font = Font(name=F, size=14, bold=True, color=WHITE)
cd['B2'] = "Anything not present in all three years is a categorisation change, not necessarily a business change."
cd['B2'].font = Font(name=F, size=10, italic=True, color="C9D1E0")

H(cd, 4, 2, "Cat1 value", 30, "left")
H(cd, 4, 3, "2023", 14)
H(cd, 4, 4, "2024", 14)
H(cd, 4, 5, "2025", 14)
H(cd, 4, 6, "Status", 40, "left")

DRIFT = [
    ("--Unapplied Payments--", 172.00, -100.00, 11920.58, "All years. 2025 balance needs clearing."),
    ("Annual Membership", 58293.50, 61640.05, 52776.95, "All years. Student-count proxy."),
    ("Birthday Parties", 45767.05, 19038.20, None, "GONE in 2025 — appears discontinued."),
    ("Gift Certificate", 310.00, 270.00, None, "GONE in 2025."),
    ("Non-Instructional time", None, None, 340.00, "NEW in 2025. Too small to be the parties successor."),
    ("Open Gym", 30059.50, 23357.15, 28128.56, "All years."),
    ("Private Lessons", 3815.71, 5645.13, 8318.25, "All years. Growing fast."),
    ("ProShop", 28031.65, 28789.86, 34975.39, "All years."),
    ("Recreational", 1309449.56, 1596000.83, 1536437.45, "All years. The core."),
    ("Staff", 2660.10, 10226.82, 3824.77, "All years. 2024 spike unexplained."),
    ("Team", 517731.45, 552485.77, 564294.62, "All years."),
    ("Tumble", None, None, 17.80, "NEW in 2025. A mis-tag — belongs under Recreational."),
]
r = 5
for name, a, b, c_, status in DRIFT:
    lc = cd.cell(row=r, column=2, value=name)
    lc.fill = lab
    lc.font = Font(name=F, size=10, bold=True, color="1F2937")
    lc.border = box
    lc.alignment = Alignment(horizontal="left", indent=1)
    for col, val in ((3, a), (4, b), (5, c_)):
        cc = cd.cell(row=r, column=col, value=val if val is not None else "absent")
        cc.fill = white if val is not None else PatternFill("solid", fgColor="F5D9D9")
        cc.font = inp if val is not None else Font(name=F, size=9, italic=True, color="8B0000")
        if val is not None:
            cc.number_format = M2
        cc.border = box
        cc.alignment = Alignment(horizontal="center")
    cc = cd.cell(row=r, column=6, value=status)
    cc.fill = calc
    cc.font = Font(name=F, size=9, color="1F2937")
    cc.border = box
    cc.alignment = Alignment(horizontal="left", indent=1)
    r += 1

r += 1
for text, font in [
    ("Cat2 drift inside Recreational", Font(name=F, size=11, bold=True, color=WHITE)),
    ("Pre-Team (Invite Only): $26,248 in 2023, then absent. Pre-team revenue moves into Rec Girls / Rec Boys via", w),
    ("Cat3 codes (Girls PT1, Boys PT1) from 2024. Girls Wings 2023-to-2024 growth is therefore overstated.", w),
    ("Boys B3 collapses from $11,108 (2023) to $10,445 (2024) to $1,230 (2025) — a level being retired.", w),
    ("Girls G4 and G5 both grow sharply in 2025 ($124,249 and $54,638) while G1 falls — students moving up the", w),
    ("ladder rather than new students entering. Consistent with the falling membership-fee count.", w),
]:
    cd.cell(row=r, column=2, value=text).font = font
    r += 1

for sheet in wbk.worksheets:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

wbk.save(OUT)
for y, idx in (("2023", 2), ("2024", 3), ("2025", 4), ("2026", 5)):
    tsum = sum(x[idx] for x in ROWS)
    print("%s sum=%.2f reported=%.2f var=%.2f" % (y, tsum, REPORTED[y], tsum - REPORTED[y]))
