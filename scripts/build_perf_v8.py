import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

OUT = sys.argv[1]

DARK, PURPLE, LABELBG, CALCBG, WHITE = "2F3641", "8B3DD7", "D6DEF5", "DCE3F2", "FFFFFF"
AMBER = "FFD479"
F = "Arial"

w    = Font(name=F, size=10, color=WHITE)
wbf  = Font(name=F, size=10, bold=True, color=WHITE)
hdrf = Font(name=F, size=9, bold=True, color=WHITE)
inp  = Font(name=F, size=9, color="0000FF")
cal  = Font(name=F, size=9, color="000000")
tot  = Font(name=F, size=9, bold=True, color="000000")

dark  = PatternFill("solid", fgColor=DARK)
purp  = PatternFill("solid", fgColor=PURPLE)
lab   = PatternFill("solid", fgColor=LABELBG)
calc  = PatternFill("solid", fgColor=CALCBG)
white = PatternFill("solid", fgColor=WHITE)
totf  = PatternFill("solid", fgColor="B9C4DC")
subf  = PatternFill("solid", fgColor="CFD8EA")
flagf = PatternFill("solid", fgColor="E8D9F5")
exclf = PatternFill("solid", fgColor="D9D9D9")
na    = PatternFill("solid", fgColor="EDEDED")

thin = Side(style="thin", color="9AA3B2")
box  = Border(left=thin, right=thin, top=thin, bottom=thin)

NUM, MONEY, M2, PCT = '#,##0;(#,##0);-', '$#,##0;($#,##0);-', '$#,##0.00;($#,##0.00);-', '0.0%;(0.0%);-'
YEARS = ["2023", "2024", "2025", "2026 YTD"]
NY = 4

# label, kind, unit, units(4) | None | "FORMULA", revenue(4)
ROWS = [
    ("Recreational -- Preschool (The Jungle)",   "E", "enrolment",  (1151, 1119, 1071, None), (343335.26, 381504.55, 367014.79, 194726.85)),
    ("Recreational -- Girls Wings",              "E", "enrolment",  (2090, 2517, 2440, None), (748491.02, 999839.46, 975278.37, 505818.42)),
    ("Recreational -- Boys Wings",               "E", "enrolment",  (277,  247,  239,  None), (86323.49,   84553.60,  85489.15,  40964.00)),
    ("Recreational -- Tumbling",                 "E", "enrolment",  (332,  433,  334,  None), (83307.09,  102217.79,  89018.18,  36851.55)),
    ("Recreational -- Workshops & Class Series", "E", "enrolment",  (1,    5,    128,  None), (36.00,        671.40,   3134.00,     25.00)),
    ("Recreational -- Unlinked Tuition",         "F", "n/a",        None,                     (21708.45,   27214.03,  16502.96,   9443.75)),
    ("Competitive -- Pre-Team",                  "E", "enrolment",  (41,   0,    0,    None), (26248.25,        0.0,       0.0,       0.0)),
    ("Competitive -- American Flyers Teams",     "E", "enrolment",  (1792, 1878, 2129, None), (517731.45, 552485.77, 564294.62, 259212.55)),
    ("Ancillary -- Open Gym",                    "U", "booking",    (2364, 1849, 2114, None), (30059.50,   23357.15,  28128.56,  20980.25)),
    ("Ancillary -- Annual Membership Fees",      "U", "membership", "FORMULA",                (58293.50,   61640.05,  52776.95,  21575.30)),
    ("Ancillary -- Birthday Parties",            "U", "party",      (112,  54,   0,    0),    (45767.05,   19038.20,       0.0,       0.0)),
    ("Ancillary -- Pro Shop",                    "U", "item sold",  (660,  789,  1042, 438),  (28031.65,   28789.86,  34975.39,  15754.09)),
    ("Ancillary -- Private Lessons",             "U", "lesson",     (63,   110,  111,  80),   (3815.71,     5645.13,   8318.25,   6055.90)),
    ("Ancillary -- Gift Certificates",           "U", "certificate",(11,   6,    0,    0),    (310.00,       270.00,       0.0,       0.0)),
    ("Ancillary -- Non-Instructional Time",      "N", "n/a",        None,                     (0.0,             0.0,     340.00,       0.0)),
    ("Other -- Tumble (stray Cat1)",             "F", "n/a",        None,                     (0.0,             0.0,      17.80,       0.0)),
    ("Review -- Unapplied Payments",             "F", "n/a",        None,                     (172.00,      -100.00,  11920.58,  12817.63)),
    ("EXCLUDE -- Staff",                         "X", "n/a",        None,                     (2660.10,    10226.82,   3824.77,   1998.03)),
]
TARGETS = [1996290.52, 2297353.81, 2241034.37, 1126223.32]

C_LABEL, C_UNIT = 2, 3
C_QTY  = [4, 5, 6, 7]
C_RPU  = [8, 9, 10, 11]
C_REV  = [12, 13, 14, 15]
C_COS  = [16, 17, 18, 19]
C_COSP = [20, 21, 22, 23]
C_GP   = [24, 25, 26, 27]
LAST = 27

GH, SH = 15, 16
TR = 17
BR = TR + len(ROWS) - 1
TOT = BR + 1
SUB = TOT + 1

wbk = Workbook()
ws = wbk.active
ws.title = "Service Performance"


def canvas(sheet, nr, nc):
    sheet.sheet_view.showGridLines = False
    for r in range(1, nr + 1):
        for c in range(1, nc + 1):
            cell = sheet.cell(row=r, column=c)
            cell.fill = dark
            cell.font = w


canvas(ws, 62, LAST + 2)
ws.column_dimensions['A'].width = 2
ws.column_dimensions['B'].width = 40
ws.column_dimensions['C'].width = 12
for c in range(4, LAST + 1):
    ws.column_dimensions[ws.cell(row=1, column=c).column_letter].width = 12


def L(col):
    return ws.cell(row=1, column=col).column_letter


ws['B1'] = "Maine Academy of Gymnastics"
ws['B1'].font = Font(name=F, size=16, bold=True, color=WHITE)
ws['B2'] = "Service Line Performance — 2023 / 2024 / 2025 / 2026 YTD"
ws['B2'].font = Font(name=F, size=11, color="C9D1E0")
ws['B3'] = ("REVENUE: Jackrabbit Revenue Summary, one report per period. All four reconcile to their stated totals with zero variance.  "
            "UNITS: Class/Event Revenue Summary (class rows, de-duplicated) and Sales Detail (Pro Shop, lessons, certificates).")
ws['B3'].font = Font(name=F, size=9, italic=True, color="C9D1E0")
ws['B4'] = ("2026 covers 1 Jan to 28 Jul only. Class-enrolment units are BLANK for 2026 because that report's enrolment window runs "
            "1/1/2025-7/28/2026 (19 months) and is not comparable. Revenue for 2026 is correct.")
ws['B4'].font = Font(name=F, size=9, bold=True, color=AMBER)
ws['B5'] = "THE UNIT DIFFERS BY ROW. Never sum the Units columns top to bottom — see the Unit column and the Service Definitions tab."
ws['B5'].font = Font(name=F, size=9, bold=True, color=AMBER)

# ---------- summary ----------
for i, y in enumerate(YEARS):
    c = ws.cell(row=7, column=4 + i, value=y)
    c.fill = purp
    c.font = hdrf
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.border = box

summary = [
    (8,  "Class enrolments (comparable rows)",     [f"={L(C_QTY[i])}{SUB}" for i in range(NY)], NUM),
    (9,  "Total Revenue",                          [f"={L(C_REV[i])}{TOT}" for i in range(NY)], M2),
    (10, "Revenue per class enrolment",            [f"=IFERROR({L(C_REV[i])}{SUB}/{L(C_QTY[i])}{SUB},0)" for i in range(NY)], MONEY),
    (12, "Cost of Sales (as a % of Revenue)",      [f"=IFERROR({L(C_COS[i])}{TOT}/{L(C_REV[i])}{TOT},0)" for i in range(NY)], PCT),
    (13, "Gross Profit",                           [f"={L(C_GP[i])}{TOT}" for i in range(NY)], M2),
    (14, "Gross Profit (as a % of Revenue)",       [f"=IFERROR({L(C_GP[i])}{TOT}/{L(C_REV[i])}{TOT},0)" for i in range(NY)], PCT),
]
for row, label, formulas, fmt in summary:
    lc = ws.cell(row=row, column=C_LABEL, value=label)
    lc.font = wbf
    lc.alignment = Alignment(horizontal="right", vertical="center")
    for i, f in enumerate(formulas):
        c = ws.cell(row=row, column=4 + i, value=f)
        c.fill = white
        c.font = tot
        c.number_format = fmt
        c.border = box
        c.alignment = Alignment(horizontal="center")

# ---------- header ----------
groups = [("Product/Service", [C_LABEL]), ("Unit", [C_UNIT]), ("# Units Sold", C_QTY),
          ("Revenue per Unit", C_RPU), ("Total Revenue", C_REV), ("Total Cost of Sales", C_COS),
          ("Cost of Sales %", C_COSP), ("Total Gross Profit", C_GP)]
for title, cols in groups:
    c1, c2 = cols[0], cols[-1]
    if len(cols) == 1:
        ws.merge_cells(start_row=GH, start_column=c1, end_row=SH, end_column=c2)
    else:
        ws.merge_cells(start_row=GH, start_column=c1, end_row=GH, end_column=c2)
    top = ws.cell(row=GH, column=c1, value=title)
    top.fill = purp
    top.font = hdrf
    top.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for c in range(c1, c2 + 1):
        for rr in (GH, SH):
            cc = ws.cell(row=rr, column=c)
            cc.fill = purp
            cc.border = box
    if len(cols) > 1:
        for i, c in enumerate(cols):
            cc = ws.cell(row=SH, column=c, value=YEARS[i])
            cc.fill = purp
            cc.font = hdrf
            cc.alignment = Alignment(horizontal="center", vertical="center")

# ---------- rows ----------
enrol_rows = []
for idx, (label, kind, unit, qty, rev) in enumerate(ROWS):
    r = TR + idx
    if kind == "E":
        enrol_rows.append(r)
    fill = {"E": lab, "U": lab, "N": lab, "F": flagf, "X": exclf}[kind]

    lc = ws.cell(row=r, column=C_LABEL, value=label)
    lc.fill = fill
    lc.font = Font(name=F, size=9, bold=True, color="1F2937")
    lc.border = box
    lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)

    uc = ws.cell(row=r, column=C_UNIT, value=unit)
    uc.fill = fill
    uc.font = Font(name=F, size=8, italic=True, color="1F2937")
    uc.border = box
    uc.alignment = Alignment(horizontal="center", vertical="center")

    for i in range(NY):
        c = ws.cell(row=r, column=C_QTY[i])
        if qty == "FORMULA":
            c.value = f"=ROUND({L(C_REV[i])}{r}/45,0)"
            c.fill = calc
            c.font = cal
        elif qty is not None and qty[i] is not None and qty[i] != 0:
            c.value = qty[i]
            c.fill = white
            c.font = inp
        else:
            c.fill = na if (qty is None or qty[i] is None) else white
            c.font = inp
        c.number_format = NUM
        c.border = box
        c.alignment = Alignment(horizontal="center")

        c = ws.cell(row=r, column=C_REV[i], value=rev[i] if rev[i] else None)
        c.fill = white
        c.font = inp
        c.number_format = M2
        c.border = box
        c.alignment = Alignment(horizontal="center")

        c = ws.cell(row=r, column=C_COSP[i])
        c.fill = white
        c.font = inp
        c.number_format = PCT
        c.border = box
        c.alignment = Alignment(horizontal="center")

        for col, formula, fmt in (
            (C_RPU[i], f"=IFERROR({L(C_REV[i])}{r}/{L(C_QTY[i])}{r},0)", MONEY),
            (C_COS[i], f"={L(C_REV[i])}{r}*{L(C_COSP[i])}{r}", M2),
            (C_GP[i],  f"={L(C_REV[i])}{r}-{L(C_COS[i])}{r}", M2),
        ):
            c = ws.cell(row=r, column=col, value=formula)
            c.fill = calc
            c.font = cal
            c.number_format = fmt
            c.border = box
            c.alignment = Alignment(horizontal="center")

# ---------- total ----------
tl = ws.cell(row=TOT, column=C_LABEL, value="TOTAL REVENUE — all lines")
tl.fill = purp
tl.font = hdrf
tl.border = box
tl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
ws.cell(row=TOT, column=C_UNIT).fill = purp
ws.cell(row=TOT, column=C_UNIT).border = box
for i in range(NY):
    for col, formula, fmt in (
        (C_REV[i], f"=SUM({L(C_REV[i])}{TR}:{L(C_REV[i])}{BR})", M2),
        (C_COS[i], f"=SUM({L(C_COS[i])}{TR}:{L(C_COS[i])}{BR})", M2),
        (C_GP[i],  f"=SUM({L(C_GP[i])}{TR}:{L(C_GP[i])}{BR})",  M2),
        (C_COSP[i], f"=IFERROR({L(C_COS[i])}{TOT}/{L(C_REV[i])}{TOT},0)", PCT),
    ):
        c = ws.cell(row=TOT, column=col, value=formula)
        c.fill = totf
        c.font = tot
        c.number_format = fmt
        c.border = box
        c.alignment = Alignment(horizontal="center")
    c = ws.cell(row=TOT, column=C_QTY[i], value="—")
    c.fill = totf
    c.font = Font(name=F, size=9, bold=True, color="7F7F7F")
    c.border = box
    c.alignment = Alignment(horizontal="center")

# ---------- enrolment subtotal ----------
sl = ws.cell(row=SUB, column=C_LABEL, value="Subtotal — class enrolments only (comparable)")
sl.fill = subf
sl.font = Font(name=F, size=9, bold=True, color="1F2937")
sl.border = box
sl.alignment = Alignment(horizontal="left", vertical="center", indent=1)
uc = ws.cell(row=SUB, column=C_UNIT, value="enrolment")
uc.fill = subf
uc.font = Font(name=F, size=8, italic=True, color="1F2937")
uc.border = box
uc.alignment = Alignment(horizontal="center")
for i in range(NY):
    qsum = "+".join(f"{L(C_QTY[i])}{r}" for r in enrol_rows)
    rsum = "+".join(f"{L(C_REV[i])}{r}" for r in enrol_rows)
    for col, formula, fmt in ((C_QTY[i], f"={qsum}", NUM),
                              (C_REV[i], f"={rsum}", M2),
                              (C_RPU[i], f"=IFERROR({L(C_REV[i])}{SUB}/{L(C_QTY[i])}{SUB},0)", MONEY)):
        c = ws.cell(row=SUB, column=col, value=formula)
        c.fill = subf
        c.font = Font(name=F, size=9, bold=True, color="1F2937")
        c.number_format = fmt
        c.border = box
        c.alignment = Alignment(horizontal="center")

# ---------- reconciliation ----------
r = SUB + 2
ws.cell(row=r, column=2, value="RECONCILIATION").font = Font(name=F, size=11, bold=True, color=WHITE)
r += 1
lc = ws.cell(row=r, column=2, value="Report stated total")
lc.font = wbf
lc.alignment = Alignment(horizontal="right")
for i in range(NY):
    c = ws.cell(row=r, column=C_REV[i], value=TARGETS[i])
    c.fill = white
    c.font = tot
    c.number_format = M2
    c.border = box
    c.alignment = Alignment(horizontal="center")
rep = r
r += 1
lc = ws.cell(row=r, column=2, value="Variance (must be zero)")
lc.font = wbf
lc.alignment = Alignment(horizontal="right")
for i in range(NY):
    letter = L(C_REV[i])
    c = ws.cell(row=r, column=C_REV[i], value=f"={letter}{TOT}-{letter}{rep}")
    c.fill = white
    c.font = tot
    c.number_format = M2
    c.border = box
    c.alignment = Alignment(horizontal="center")
r += 2

for text, font in [
    ("HOW TO USE THIS SHEET", Font(name=F, size=11, bold=True, color=WHITE)),
    ("Fill the BLUE cells: # Units Sold, Total Revenue, Cost of Sales %. Everything else calculates.", w),
    ("Grey units cells mean no count is available for that row and period. Membership units are a formula (revenue / $45).", w),
    ("", w),
    ("UNITS ARE NOT COMPARABLE ACROSS ROWS. A class enrolment, an Open Gym booking, a party and a leotard are", Font(name=F, size=10, bold=True, color=AMBER)),
    ("different things. The TOTAL row shows a dash for units; the subtotal beneath covers class-enrolment rows only.", w),
    ("", w),
    ("Cost of Sales is not in Jackrabbit. For class rows it is coach labour (hourly rate x class hours) from payroll.", Font(name=F, size=10, bold=True, color=AMBER)),
    ("Pro Shop is the exception — its cost of sales is inventory purchase cost. Treat that row separately.", w),
    ("", w),
    ("See 'Store Detail' for Falmouth Rec, visiting-team rental, flex passes and late fees — real revenue that could not", w),
    ("be placed on this grid because its Jackrabbit revenue category could not be verified.", w),
    ("", w),
    ("See 'Service Definitions' for what each row contains, what one unit means, and the open data-quality items.", Font(name=F, size=10, italic=True, color="C9D1E0")),
]:
    ws.cell(row=r, column=2, value=text).font = font
    r += 1

ws.freeze_panes = "D17"

# ==================== Store Detail ====================
st = wbk.create_sheet("Store Detail")
canvas(st, 58, 12)
st.column_dimensions['A'].width = 2
st.column_dimensions['B'].width = 34
st.column_dimensions['C'].width = 12
for col in ('D', 'E', 'F', 'G', 'H', 'I', 'J', 'K'):
    st.column_dimensions[col].width = 12
st.column_dimensions['L'].width = 60

st['B1'] = "Store Detail — item-level lines not on the main grid"
st['B1'].font = Font(name=F, size=15, bold=True, color=WHITE)
st['B2'] = "Source: Jackrabbit Sales Detail reports, 2023-2026 YTD. Quantities and revenue are as reported by the Store module."
st['B2'].font = Font(name=F, size=10, italic=True, color="C9D1E0")
st['B3'] = ("These are on the STORE basis and are NOT added into the main grid total. Store revenue does not tie exactly to the "
            "Revenue Summary categories, so double-counting would result.")
st['B3'].font = Font(name=F, size=9, bold=True, color=AMBER)

H1 = 5
st.merge_cells(start_row=H1, start_column=4, end_row=H1, end_column=7)
st.merge_cells(start_row=H1, start_column=8, end_row=H1, end_column=11)
for col, txt in ((2, "Store line"), (3, "Unit"), (4, "Quantity"), (8, "Revenue"), (12, "Notes")):
    c = st.cell(row=H1, column=col, value=txt)
    c.fill = purp
    c.font = hdrf
    c.border = box
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for c in list(range(4, 12)):
    st.cell(row=H1, column=c).fill = purp
    st.cell(row=H1, column=c).border = box
st.merge_cells(start_row=H1, start_column=2, end_row=H1 + 1, end_column=2)
st.merge_cells(start_row=H1, start_column=3, end_row=H1 + 1, end_column=3)
st.merge_cells(start_row=H1, start_column=12, end_row=H1 + 1, end_column=12)
for i, y in enumerate(YEARS):
    for base in (4, 8):
        c = st.cell(row=H1 + 1, column=base + i, value=y)
        c.fill = purp
        c.font = hdrf
        c.border = box
        c.alignment = Alignment(horizontal="center")

STORE = [
    ("Falmouth Rec (contract programme)", "billing", (64, 58, 30, 0), (8258.00, 8201.20, 4242.00, 0.0),
     "Bulk-billed in one or two lines a year. Declining then gone in 2026. Looks like a contract with Falmouth's "
     "recreation department — a B2B channel, not retail. Does NOT appear anywhere in the Revenue Summary by name, "
     "so its Cat1 could not be verified. Most likely inside Recreational Unlinked Tuition (which is $21,708 in 2023)."),
    ("Visiting team gym rental", "rental", (25, 15, 43, 8), (2197.00, 640.00, 1620.00, 315.00),
     "Facility hire to visiting teams. Near-zero marginal cost. Revenue Summary shows only $1,192 (2023) and $80 "
     "(2025) under Cat1 Team by this name, so most of it posts elsewhere unnamed."),
    ("Summer Flex Passes", "pass", (0, 0, 0, 10), (0.0, 0.0, 0.0, 2260.50),
     "NEW IN 2026 — Red, Blue, White and First Flight variants. May be cannibalising Open Gym. Worth watching."),
    ("Late fees", "charge", (8, 7, 24, 46), (200.00, 175.00, 575.00, 1198.84),
     "Rising sharply: 8 to 46 charges. Consistent with the growing unapplied-payments balance. Collections are drifting."),
    ("Xcel clinic", "place", (0, 33, 0, 0), (0.0, 825.00, 0.0, 0.0),
     "One-off Xcel Regionals prep clinic in 2024."),
    ("Birthday parties — base", "party", (112, 54, 0, 0), (27660.00, 13500.00, 0.0, 0.0),
     "Cross-validates the Class/Event report, which independently showed 112 parties in 2023. Discontinued after 2024."),
    ("Birthday parties — extra child", "child", (219, 138, 0, 0), (2190.00, 1380.00, 0.0, 0.0),
     "Roughly 2 extra children per party, at $10 each."),
    ("Birthday parties — tips", "tip", (82, 40, 0, 0), (4554.00, 1900.00, 0.0, 0.0),
     "Tips collected on parties. Averaged about $41 per party in 2023."),
    ("Team apparel (all SKUs)", "item sold", (None, None, None, None), (11214.50, 13468.87, 5329.20, 1480.41),
     "Team Leos, Champion jackets and leggings, backpacks, competition tanks. Posts under Cat1 TEAM, not ProShop — "
     "which is why Store merchandise does not tie to the Pro Shop row. Already inside the Teams revenue line."),
    ("Retail merchandise (excl. team kit)", "item sold", (660, 789, 1042, 438), (23534.65, 27837.17, 36037.48, 15795.63),
     "This is the basis for the Pro Shop units on the main grid. Units up 58% over two years while average item "
     "value fell from $35.66 to $34.59 — volume growth partly from Summer Clearance $10/$15 SKUs added in 2025."),
    ("of which: Destira leotards", "item sold", (503, 527, 607, 266), (19166.30, 20830.17, 26289.30, 10502.72),
     "The core uniform SKU. $40 plus 5.5% Maine sales tax = the $42.20 that recurs across dozens of classes. "
     "Biggest single retail line by far, and it grows every year."),
    ("Annual membership (via Store)", "membership", (813, 871, 851, 345), (36225.00, 39105.00, 37980.00, 15525.00),
     "The $45 fee sold through the Store. Lower than the Cat1 Annual Membership total ($58,294 in 2023), so "
     "membership is also billed by other routes. Do not use this as the student count."),
]
r = H1 + 2
for name, unit, qty, rev, note in STORE:
    lc = st.cell(row=r, column=2, value=name)
    lc.fill = lab
    lc.font = Font(name=F, size=9, bold=True, color="1F2937")
    lc.border = box
    lc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    uc = st.cell(row=r, column=3, value=unit)
    uc.fill = lab
    uc.font = Font(name=F, size=8, italic=True, color="1F2937")
    uc.border = box
    uc.alignment = Alignment(horizontal="center", vertical="center")
    for i in range(NY):
        c = st.cell(row=r, column=4 + i, value=qty[i] if qty[i] else None)
        c.fill = white if qty[i] is not None else na
        c.font = cal
        c.number_format = NUM
        c.border = box
        c.alignment = Alignment(horizontal="center")
        c = st.cell(row=r, column=8 + i, value=rev[i] if rev[i] else None)
        c.fill = white
        c.font = cal
        c.number_format = M2
        c.border = box
        c.alignment = Alignment(horizontal="center")
    nc = st.cell(row=r, column=12, value=note)
    nc.fill = calc
    nc.font = Font(name=F, size=8, color="1F2937")
    nc.border = box
    nc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    st.row_dimensions[r].height = 60
    r += 1

r += 1
for text, font in [
    ("WHY FALMOUTH REC AND VISITING-TEAM RENTAL ARE NOT ON THE MAIN GRID", Font(name=F, size=11, bold=True, color=AMBER)),
    ("Both are real revenue with genuinely different economics, and both should eventually be service rows. The blocker is", w),
    ("attribution: neither appears by name in the Revenue Summary, so their Jackrabbit revenue category is unknown. Adding", w),
    ("them to the grid would mean reducing some other row by an unverified amount, which would break the reconciliation.", w),
    ("", w),
    ("THE ONE LOOKUP THAT FIXES THIS: in Jackrabbit go to Store > item setup and check which revenue category the", Font(name=F, size=10, bold=True, color=AMBER)),
    ("FALREC and VTM items are mapped to. Once that is known both can be promoted to the main grid properly, and it may", w),
    ("also explain a large share of the Recreational Unlinked Tuition bucket — Falmouth Rec alone is $8,258 in 2023,", w),
    ("about 38% of that year's unlinked total.", w),
    ("", w),
    ("Source: Jackrabbit Sales Detail 2023, 2024, 2025 and 2026 YTD (to 27 July 2026), 81 distinct item numbers.", Font(name=F, size=9, italic=True, color="C9D1E0")),
]:
    st.cell(row=r, column=2, value=text).font = font
    r += 1


# ==================== Service Definitions ====================
sd = wbk.create_sheet("Service Definitions")
canvas(sd, 48, 7)
sd.column_dimensions['A'].width = 2
for col, wd in (('B', 36), ('C', 12), ('D', 34), ('E', 28), ('F', 58)):
    sd.column_dimensions[col].width = wd

sd['B1'] = "Service Definitions"
sd['B1'].font = Font(name=F, size=15, bold=True, color=WHITE)
sd['B2'] = "Rows come from Category 1 / Category 2 in the Jackrabbit Revenue Summary. The unit differs by row — that is deliberate."
sd['B2'].font = Font(name=F, size=10, italic=True, color="C9D1E0")

for i, h in enumerate(["Service Row", "Unit", "Jackrabbit source", "What one unit is", "Notes, 2023-2026 YTD"]):
    c = sd.cell(row=4, column=2 + i, value=h)
    c.fill = purp
    c.font = hdrf
    c.border = box
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True, indent=1)

DEFS = [
    ("Recreational -- Preschool (The Jungle)", "enrolment", "Cat1 Recreational + Cat2 Pre-School or Parent-Child",
     "One child in one preschool class for one session.",
     "1,151 / 1,119 / 1,071 units. $343k / $382k / $367k / $195k. Units down 7% over two years. The funnel feeder for every other line, so the decline matters more than its size."),
    ("Recreational -- Girls Wings", "enrolment", "Cat1 Recreational + Cat2 Rec Girls",
     "One girl in one class for one session. A student in two classes counts twice.",
     "2,090 / 2,517 / 2,440 units. $748k / $1,000k / $975k / $506k. 43.5% of 2025 revenue - the largest line. Absorbs pre-team girls from 2024, so 2024 growth is slightly overstated."),
    ("Recreational -- Boys Wings", "enrolment", "Cat1 Recreational + Cat2 Rec Boys",
     "One boy in one class for one session.",
     "277 / 247 / 239 units with revenue flat at ~$86k. Three consecutive unit declines - charging more to fewer students. Only 21 class records vs 163 for girls, at 51.6% fill."),
    ("Recreational -- Tumbling", "enrolment", "Cat1 Recreational + Cat2 Tumble (Combined)",
     "One student in one tumbling class for one session.",
     "332 / 433 / 334 units. Spiked then fell back. Best fill rate in the current class list at 81.6%, so there may be room to add capacity."),
    ("Recreational -- Workshops & Class Series", "enrolment", "Cat1 Recreational, Cat2 BLANK",
     "One student in one workshop or summer series.",
     "128 units / $3,134 in 2025 - a new short-format offering (3 Summer Tumble series, 3 Summer Wings series, 2 Bar Workshops). ~$24 per unit because priced per session. TAGGING NOTE: no Category 2 set; setting it moves ~$1,364 to Tumbling and ~$1,770 to Girls Wings."),
    ("Recreational -- Unlinked Tuition", "n/a", "Cat1 Recreational, activity 'No Activity Name'",
     "No unit - no class record exists to enrol against.",
     "$21,708 / $27,214 / $16,503 / $9,444 - DOWN 54% from the 2024 peak, so hygiene is improving. Note Falmouth Rec ($8,258 in 2023) may sit inside this bucket - see Store Detail."),
    ("Competitive -- Pre-Team", "enrolment", "Cat1 Recreational + Cat2 Pre-Team (Invite Only)",
     "One athlete in one pre-team class.",
     "41 units / $26,248 in 2023 at $640 per unit - the HIGHEST unit value in the business, roughly double the rec average. The Cat2 exists in 2023 only; from 2024 these athletes sit inside the Wings rows."),
    ("Competitive -- American Flyers Teams", "enrolment", "Cat1 Team",
     "One athlete in one team activity. Counts billing groups AND meet entries, so it double-counts athletes.",
     "1,792 / 1,878 / 2,129 units. Units up 13.4% in 2025 while revenue per unit FELL from $294 to $265 - more meet participation, which is pass-through at near-zero margin. In 2023 only $377,627 was tuition; $140,104 was meet entry fees. Also contains team apparel (~$11k)."),
    ("Ancillary -- Open Gym", "booking", "Cat1 Open Gym",
     "One prepaid session booking.",
     "2,364 / 1,849 / 2,114 bookings at a flat ~$13. Highest unit count, lowest unit value. Marginal cost is supervision only, so margin should be well above the tuition rows. Watch the new 2026 Flex Passes for cannibalisation."),
    ("Ancillary -- Annual Membership Fees", "membership", "Cat1 Annual Membership",
     "One membership - the $45 annual fee, one per student per year.",
     "$58,294 / $61,640 / $52,777 / $21,575. Units are a FORMULA (revenue / 45). Essentially 100% margin. CAUTION: implies a 14% fall in students in 2025, which the class enrolment data contradicts - treat as a billing-timing artefact, NOT a headcount measure."),
    ("Ancillary -- Birthday Parties", "party", "Cat1 Birthday Parties",
     "One party booked.",
     "112 / 54 parties then nil. $45,767 / $19,038. Wound down to nothing and the category disappears in 2025. Was 2.3% of revenue with almost no coach cost. Store data adds detail: ~2 extra children per party at $10, plus ~$41 of tips per party."),
    ("Ancillary -- Pro Shop", "item sold", "Cat1 ProShop",
     "One item sold, from the Sales Detail report.",
     "660 / 789 / 1,042 / 438 items. $28,032 / $28,790 / $34,975 / $15,754. Units up 58% while average item value fell $35.66 to $34.59 - volume growth partly from Summer Clearance SKUs. Destira leotard at $40 + 5.5% tax = $42.20 is the core SKU. THIS IS THE ONE ROW WHERE COST OF SALES IS INVENTORY COST, NOT COACH LABOUR. Team apparel is excluded - it posts under Cat1 Team."),
    ("Ancillary -- Private Lessons", "lesson", "Cat1 Private Lessons",
     "One lesson, from the Sales Detail report (1hr, 45min, 30min and semi-private).",
     "63 / 110 / 111 / 80 lessons. $3,816 / $5,645 / $8,318 / $6,056 - up 118% over two years. Store counts cover most but not all of the revenue, so units are indicative. Premium one-to-one delivery, constrained by coach time."),
    ("Ancillary -- Gift Certificates", "certificate", "Cat1 Gift Certificate",
     "One certificate sold.",
     "11 / 6 certificates. $310 / $270, then discontinued alongside Birthday Parties. Store revenue matches the Cat1 total exactly, which is a useful cross-check that Store items do map to categories."),
    ("Ancillary -- Non-Instructional Time", "n/a", "Cat1 Non-Instructional time",
     "No meaningful unit.",
     "$340 in 2025 only. A new category - worth asking what it is before it grows."),
    ("Other -- Tumble (stray Cat1)", "n/a", "Cat1 Tumble",
     "No unit.",
     "$17.80 in 2025. A mis-tag; belongs under Recreational + Tumble (Combined). Trivial, but a symptom of the same category discipline issue."),
    ("Review -- Unapplied Payments", "n/a", "Cat1 --Unapplied Payments--",
     "No unit - cash received but not applied to any fee.",
     "$172 / -$100 / $11,921 / $12,818. GETTING WORSE, and consistent with late fees rising from 8 to 46 charges. Real money sitting unallocated on family accounts. Use Family Balance Summary and Aged Accounts Details to clear it."),
    ("EXCLUDE -- Staff", "n/a", "Cat1 Staff",
     "No unit - internal transactions.",
     "$2,660 / $10,227 / $3,825 / $1,998. Not customer revenue; exclude from any operating view. The 2024 spike needs an explanation."),
]
r = 5
for name, unit, src, unitdesc, notes in DEFS:
    for i, v in enumerate([name, unit, src, unitdesc, notes]):
        c = sd.cell(row=r, column=2 + i, value=v)
        if i == 0:
            c.fill = lab
            c.font = Font(name=F, size=9, bold=True, color="1F2937")
        elif i == 1:
            c.fill = lab
            c.font = Font(name=F, size=8, italic=True, color="1F2937")
        else:
            c.fill = calc
            c.font = Font(name=F, size=8, color="1F2937")
        c.border = box
        c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True, indent=1)
    sd.row_dimensions[r].height = 96
    r += 1

r += 1
for text, font in [
    ("Why the unit differs by row", Font(name=F, size=11, bold=True, color=AMBER)),
    ("A gym sells several different things. Classes sell enrolments, Open Gym sells session bookings, the shop sells items, and", w),
    ("the membership fee sells one membership per student per year. Forcing all of those into one 'transactions' count would", w),
    ("measure billing cadence rather than demand - a student enrolled ten months generates ten charges and one enrolment.", w),
    ("", w),
    ("Consequence: never sum the Units columns, and compare Revenue per Unit only between rows sharing a unit.", w),
    ("", w),
    ("Sources: Revenue Summary (revenue), Class/Event Revenue Summary (class units, de-duplicated on Class + Session +", Font(name=F, size=9, italic=True, color="C9D1E0")),
    ("Category 3), Sales Detail (Pro Shop, lessons, certificates). All run 28 July 2026. All four periods reconcile exactly.", Font(name=F, size=9, italic=True, color="C9D1E0")),
]:
    sd.cell(row=r, column=2, value=text).font = font
    r += 1


for sheet in wbk.worksheets:
    sheet.page_setup.orientation = "landscape"
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = 0
    sheet.sheet_properties.pageSetUpPr.fitToPage = True

wbk.save(OUT)
for i, y in enumerate(YEARS):
    s = sum(x[4][i] for x in ROWS)
    print("%-9s rev=%14.2f target=%14.2f var=%.2f" % (y, s, TARGETS[i], s - TARGETS[i]))
print("grid rows:", len(ROWS), "| store rows:", len(STORE), "| enrolment rows:", sum(1 for x in ROWS if x[1] == "E"))
